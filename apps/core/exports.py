"""通用 Excel 导出（M7-7 / #10，M12 美化）。

- xlsx_response(report_name, headers, rows, ...)：输出带标题/编制单位/期间/导出日期 + 表格样式的 xlsx；
  并按统一规则生成下载文件名。
- resolve_cell(obj, accessor)：accessor 支持属性名 / "a__b" 跨级 / callable / get_x_display。
"""

import re
from datetime import date, datetime
from decimal import Decimal
from urllib.parse import quote

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# 跨 2 家及以上公司的导出，抬头统一用集团名
GROUP_HEADER = "鸿威达新材料集团公司"

_HEADER_FILL = PatternFill("solid", fgColor="E8EEF7")
_TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
_THIN = Side(style="thin", color="C0C0C0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right")
_LEFT = Alignment(horizontal="left")


def resolve_cell(obj, accessor):
    """取值：callable(obj) / 字段路径 'supplier__name' / 'get_status_display' 自动调用。"""
    if callable(accessor):
        return accessor(obj)
    cur = obj
    for part in accessor.split("__"):
        if cur is None:
            return ""
        cur = getattr(cur, part, None)
        if callable(cur):  # 如 get_status_display
            cur = cur()
    return cur


def _norm(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "是" if v else "否"
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float, str)):
        return v
    return str(v)  # 模型实例 / 其它对象统一转字符串，避免 openpyxl 写入报错


def _is_num(v):
    return isinstance(v, (int, float, Decimal)) and not isinstance(v, bool)


def _disp_width(s):
    """估算列宽：中文按 2、其它按 1。"""
    return sum(2 if ord(ch) > 127 else 1 for ch in str(s))


def _safe(s):
    s = re.sub(r'[\\/:*?"<>|\n\r\t]+', "", str(s or "")).strip()
    return s.replace(" ", "")


def xlsx_response(report_name, headers, rows, *, company=None, period=None,
                  sheet_title=None, generated=None, extra_meta=None):
    """report_name 报表名（作标题）；company 编制单位（Company 或 str）；
    period (dfrom, dto) 期间；generated 导出日期（默认今天）。
    文件名规则：{报表名}_{编制单位}_{起}-{止}_{导出日}.xlsx（缺省部分省略）。"""
    # 抬头：单公司用其法定全称，跨 2 家及以上(company=None)用集团名
    header_label = (getattr(company, "header_name", None)
                    or (str(company) if company is not None else GROUP_HEADER))
    # 文件名用短名，避免过长
    company_short = (getattr(company, "short_name", None) or getattr(company, "name", None)
                     or (str(company) if company is not None else "集团"))
    gen = generated or timezone.localdate()
    dfrom = dto = None
    if period:
        dfrom, dto = period

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or report_name or "数据")[:31]
    ncols = max(1, len(headers))
    last_col = get_column_letter(ncols)

    r = 1
    # 抬头（公司法定全称 / 集团名）
    ws.merge_cells(f"A{r}:{last_col}{r}")
    hc = ws.cell(r, 1, header_label)
    hc.font = Font(bold=True, size=15)
    hc.alignment = _CENTER
    ws.row_dimensions[r].height = 30
    r += 1
    # 报表名
    ws.merge_cells(f"A{r}:{last_col}{r}")
    cell = ws.cell(r, 1, report_name)
    cell.font = Font(bold=True, size=12)
    cell.alignment = _CENTER
    ws.row_dimensions[r].height = 20
    r += 1
    # 元信息行：期间 / 导出日期（编制单位已在抬头体现）
    meta = []
    if dfrom or dto:
        meta.append(f"期间：{_norm(dfrom) or '起初'} ~ {_norm(dto) or '至今'}")
    for m in (extra_meta or []):
        meta.append(str(m))
    meta.append(f"导出日期：{_norm(gen)}")
    ws.merge_cells(f"A{r}:{last_col}{r}")
    mcell = ws.cell(r, 1, "　".join(meta))
    mcell.font = Font(size=9, color="808080")
    mcell.alignment = _LEFT
    r += 1

    # 左上角嵌入 logo（缺 Pillow / 文件不存在则跳过，绝不影响导出）
    try:
        from django.conf import settings
        from openpyxl.drawing.image import Image as XLImage
        logo_path = settings.BASE_DIR / "static" / "img" / "logo.png"
        if logo_path.exists():
            img = XLImage(str(logo_path))
            img.height, img.width = 38, round(38 * img.width / img.height)
            ws.add_image(img, "A1")
    except Exception:
        pass

    # 表头
    header_row = r
    for j, h in enumerate(headers, 1):
        c = ws.cell(header_row, j, h)
        c.font = Font(bold=True)
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = _CENTER
    r += 1

    # 数据
    for row in rows:
        is_total = bool(row) and str(row[0]).strip() in ("合计", "总计")
        for j, v in enumerate(row, 1):
            c = ws.cell(r, j, _norm(v))
            c.border = _BORDER
            if _is_num(v):
                c.alignment = _RIGHT
                c.number_format = "#,##0.00"   # 两位小数 + 千分位
            if is_total:
                c.font = Font(bold=True)
                c.fill = _TOTAL_FILL
        r += 1

    # 列宽
    for j, h in enumerate(headers, 1):
        w = _disp_width(h)
        for row in rows:
            if j - 1 < len(row):
                w = max(w, _disp_width(_norm(row[j - 1])))
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 9), 42)

    ws.freeze_panes = ws.cell(header_row + 1, 1)  # 冻结标题/表头

    # 文件名：{报表名}_{编制单位}_{起}-{止}_{导出日}.xlsx
    def _ymd(d):
        return d.strftime("%Y%m%d") if hasattr(d, "strftime") else ""
    parts = [_safe(report_name)]
    if company_short:
        parts.append(_safe(company_short))
    if dfrom or dto:
        parts.append(f"{_ymd(dfrom)}-{_ymd(dto)}")
    parts.append(_ymd(gen) or _safe(_norm(gen)))
    fname = "_".join(p for p in parts if p) + ".xlsx"

    resp = HttpResponse(content_type=XLSX)
    resp["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(fname)}"
    wb.save(resp)
    return resp
