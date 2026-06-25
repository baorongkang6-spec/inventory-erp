"""资金往来测试：采购发票含税换算与应付产生。"""

from datetime import date
from decimal import Decimal

from django.test import TestCase

from apps.core.models import Company
from apps.finance.models import (  # noqa: F401
    BankAccount,
    BankJournal,
    NotePayable,
    NoteReceivable,
    Payment,
    PurchaseInvoice,
    Receipt,
    SalesInvoice,
)
from apps.finance.services import (
    SettlementError,
    allocate_payment,
    allocate_receipt,
    compute_tax,
    create_payment,
    create_purchase_invoice,
    create_receipt,
    create_sales_invoice,
)
from apps.masterdata.models import Customer, Product, Supplier


class PurchaseInvoiceTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.p = Product.objects.create(company=cls.c1, code="P1", name="货A")

    def test_compute_tax(self):
        tax, taxed = compute_tax(Decimal("1000"), Decimal("0.13"))
        self.assertEqual(tax, Decimal("130.00"))
        self.assertEqual(taxed, Decimal("1130.00"))

    def test_create_invoice_produces_payable(self):
        inv = create_purchase_invoice(
            company=self.c1, user=None, doc_date=date(2026, 6, 5), supplier=self.sup,
            invoice_no="FP001",
            lines=[
                {"product": self.p, "description": "货A", "amount_untaxed": Decimal("1000"),
                 "tax_rate": Decimal("0.13")},
                {"product": None, "description": "运费", "amount_untaxed": Decimal("100"),
                 "tax_rate": Decimal("0.09")},
            ],
        )
        self.assertEqual(inv.doc_no, "CGF-C1-20260605-001")
        self.assertEqual(inv.amount_untaxed, Decimal("1100.00"))
        self.assertEqual(inv.tax_amount, Decimal("139.00"))   # 130 + 9
        self.assertEqual(inv.amount_taxed, Decimal("1239.00"))
        self.assertEqual(inv.settled_amount, Decimal("0.00"))
        self.assertEqual(inv.outstanding, Decimal("1239.00"))  # 应付余额
        self.assertEqual(inv.lines.count(), 2)

    def test_manual_tax_override_for_rounding_diff(self):
        # 录入税额/含税金额时优先采用（允许尾差手工微调）
        inv = create_purchase_invoice(
            company=self.c1, user=None, doc_date=date(2026, 6, 6), supplier=self.sup,
            lines=[
                {"product": self.p, "description": "货A", "amount_untaxed": Decimal("999.99"),
                 "tax_rate": Decimal("0.13"),
                 "tax_amount": Decimal("130.00"), "amount_taxed": Decimal("1129.99")},
            ],
        )
        ln = inv.lines.first()
        self.assertEqual(ln.tax_amount, Decimal("130.00"))      # 非自动算的 129.9987→130.00
        self.assertEqual(ln.amount_taxed, Decimal("1129.99"))
        self.assertEqual(inv.tax_amount, Decimal("130.00"))
        self.assertEqual(inv.amount_taxed, Decimal("1129.99"))


class PaymentTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户")

    def test_payment_auto_generates_bank_journal(self):
        pay = create_payment(
            company=self.c1, user=None, doc_date=date(2026, 6, 5),
            bank_account=self.acc, supplier=self.sup, amount=Decimal("500"), summary="付货款",
        )
        self.assertEqual(pay.doc_no, "FK-C1-20260605-001")
        self.assertEqual(pay.amount, Decimal("500.00"))
        self.assertEqual(pay.unallocated, Decimal("500.00"))
        # 自动生成一条支出日记账
        self.assertIsNotNone(pay.bank_journal)
        j = pay.bank_journal
        self.assertEqual(j.direction, BankJournal.Direction.OUT)
        self.assertEqual(j.amount, Decimal("500.00"))
        self.assertEqual(j.signed_amount, Decimal("-500.00"))
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 1)


class AllocationTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户")
        cls.p = Product.objects.create(company=cls.c1, code="P1", name="货A")

    def _invoice(self, untaxed, rate="0.13"):
        return create_purchase_invoice(
            company=self.c1, user=None, doc_date=date(2026, 6, 5), supplier=self.sup,
            lines=[{"product": self.p, "description": "", "amount_untaxed": Decimal(untaxed),
                    "tax_rate": Decimal(rate)}],
        )

    def _payment(self, amount):
        return create_payment(company=self.c1, user=None, doc_date=date(2026, 6, 5),
                              bank_account=self.acc, supplier=self.sup, amount=Decimal(amount))

    def test_partial_allocation(self):
        inv = self._invoice("1000")           # 含税 1130
        pay = self._payment("500")
        allocate_payment(payment=pay, allocations=[{"invoice": inv, "amount": Decimal("500")}])
        inv.refresh_from_db(); pay.refresh_from_db()
        self.assertEqual(inv.settled_amount, Decimal("500.00"))
        self.assertEqual(inv.outstanding, Decimal("630.00"))
        self.assertEqual(pay.settled_amount, Decimal("500.00"))
        self.assertEqual(pay.unallocated, Decimal("0.00"))

    def test_one_payment_multiple_invoices(self):
        a = self._invoice("100")   # 113
        b = self._invoice("200")   # 226
        pay = self._payment("339")
        allocate_payment(payment=pay, allocations=[
            {"invoice": a, "amount": Decimal("113")},
            {"invoice": b, "amount": Decimal("226")},
        ])
        a.refresh_from_db(); b.refresh_from_db(); pay.refresh_from_db()
        self.assertEqual(a.outstanding, Decimal("0.00"))
        self.assertEqual(b.outstanding, Decimal("0.00"))
        self.assertEqual(pay.unallocated, Decimal("0.00"))

    def test_over_invoice_outstanding_allowed(self):
        # 允许核销超过发票未核销额 → 发票余额变负（预付/多付）
        inv = self._invoice("100")  # 113
        pay = self._payment("500")
        allocate_payment(payment=pay, allocations=[{"invoice": inv, "amount": Decimal("200")}])
        inv.refresh_from_db(); pay.refresh_from_db()
        self.assertEqual(inv.settled_amount, Decimal("200.00"))
        self.assertEqual(inv.outstanding, Decimal("-87.00"))   # 113 - 200
        self.assertEqual(pay.settled_amount, Decimal("200.00"))

    def test_over_payment_balance_rejected(self):
        a = self._invoice("100")  # 113
        b = self._invoice("100")  # 113
        pay = self._payment("150")
        with self.assertRaises(SettlementError):
            allocate_payment(payment=pay, allocations=[
                {"invoice": a, "amount": Decimal("113")},
                {"invoice": b, "amount": Decimal("113")},  # 合计 226 > 付款 150
            ])
        a.refresh_from_db(); pay.refresh_from_db()
        self.assertEqual(a.settled_amount, Decimal("0.00"))  # 整体回滚
        self.assertEqual(pay.settled_amount, Decimal("0.00"))


class SalesSideTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户")
        cls.p = Product.objects.create(company=cls.c1, code="P1", name="货A")

    def test_sales_invoice_produces_receivable(self):
        inv = create_sales_invoice(
            company=self.c1, user=None, doc_date=date(2026, 6, 5), customer=self.cust,
            lines=[{"product": self.p, "description": "", "amount_untaxed": Decimal("1000"),
                    "tax_rate": Decimal("0.13")}],
        )
        self.assertEqual(inv.doc_no, "XSF-C1-20260605-001")
        self.assertEqual(inv.amount_taxed, Decimal("1130.00"))
        self.assertEqual(inv.outstanding, Decimal("1130.00"))

    def test_receipt_auto_journal_and_allocate(self):
        inv = create_sales_invoice(
            company=self.c1, user=None, doc_date=date(2026, 6, 5), customer=self.cust,
            lines=[{"product": self.p, "description": "", "amount_untaxed": Decimal("1000"),
                    "tax_rate": Decimal("0.13")}],
        )
        rec = create_receipt(company=self.c1, user=None, doc_date=date(2026, 6, 5),
                             bank_account=self.acc, customer=self.cust, amount=Decimal("1130"))
        self.assertEqual(rec.doc_no, "SK-C1-20260605-001")
        # 自动生成收入日记账
        self.assertEqual(rec.bank_journal.direction, BankJournal.Direction.IN)
        self.assertEqual(rec.bank_journal.signed_amount, Decimal("1130.00"))
        # 核销
        allocate_receipt(receipt=rec, allocations=[{"invoice": inv, "amount": Decimal("1130")}])
        inv.refresh_from_db(); rec.refresh_from_db()
        self.assertEqual(inv.outstanding, Decimal("0.00"))
        self.assertEqual(rec.unallocated, Decimal("0.00"))

    def test_receipt_over_allocate_allowed(self):
        # 允许核销超过发票未核销额 → 应收变负（预收/多收）
        inv = create_sales_invoice(
            company=self.c1, user=None, doc_date=date(2026, 6, 5), customer=self.cust,
            lines=[{"product": self.p, "description": "", "amount_untaxed": Decimal("100"),
                    "tax_rate": Decimal("0.13")}],
        )  # 113
        rec = create_receipt(company=self.c1, user=None, doc_date=date(2026, 6, 5),
                             bank_account=self.acc, customer=self.cust, amount=Decimal("500"))
        allocate_receipt(receipt=rec, allocations=[{"invoice": inv, "amount": Decimal("200")}])
        inv.refresh_from_db()
        self.assertEqual(inv.settled_amount, Decimal("200.00"))
        self.assertEqual(inv.outstanding, Decimal("-87.00"))


class BankJournalExcelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户",
                                             opening_balance=Decimal("1000"))

    def test_export_then_parse_roundtrip(self):
        from apps.finance.excel import export_bank_journal, parse_bank_journal_xlsx
        from apps.finance.views import _journal_rows
        from io import BytesIO
        create_payment(company=self.c1, user=None, doc_date=date(2026, 6, 5),
                       bank_account=self.acc, supplier=self.sup, amount=Decimal("300"), summary="付款A")
        _, rows, closing = _journal_rows(self.c1, self.acc)
        self.assertEqual(closing, Decimal("700.00"))  # 1000 - 300
        content = export_bank_journal(self.acc, rows)
        parsed, errors = parse_bank_journal_xlsx(BytesIO(content))
        self.assertEqual(errors, [])
        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0]["direction"], "out")
        self.assertEqual(parsed[0]["amount"], Decimal("300"))

    def test_export_has_opening_total_closing_and_reimports_clean(self):
        from io import BytesIO
        from openpyxl import load_workbook
        from apps.finance.excel import export_bank_journal, parse_bank_journal_xlsx
        from apps.finance.views import _journal_rows
        create_payment(company=self.c1, user=None, doc_date=date(2026, 6, 5),
                       bank_account=self.acc, supplier=self.sup, amount=Decimal("300"), summary="付A")
        create_payment(company=self.c1, user=None, doc_date=date(2026, 6, 7),
                       bank_account=self.acc, supplier=self.sup, amount=Decimal("200"), summary="付C")
        opening, rows, closing = _journal_rows(self.c1, self.acc, date(2026, 6, 1), date(2026, 6, 30))
        content = export_bank_journal(self.acc, rows, opening=opening, closing=closing,
                                      date_from=date(2026, 6, 1), date_to=date(2026, 6, 30))
        ws = load_workbook(BytesIO(content)).active
        grid = list(ws.iter_rows(values_only=True))
        labels = {r[0] for r in grid}
        self.assertIn("期初余额", labels)
        self.assertIn("本期合计", labels)
        self.assertIn("期末余额", labels)
        # 期初 1000、合计支出 500、期末 500
        opening_row = next(r for r in grid if r[0] == "期初余额")
        total_row = next(r for r in grid if r[0] == "本期合计")
        closing_row = next(r for r in grid if r[0] == "期末余额")
        self.assertEqual(opening_row[6], 1000)
        self.assertEqual(total_row[4], 500)   # 支出合计
        self.assertEqual(closing_row[6], 500)
        # 再导入：汇总行被跳过，只解析出 2 条真实流水、无错误
        parsed, errors = parse_bank_journal_xlsx(BytesIO(content))
        self.assertEqual(errors, [])
        self.assertEqual(len(parsed), 2)

    def test_parse_skips_header_and_blank(self):
        from openpyxl import Workbook
        from io import BytesIO
        from apps.finance.excel import parse_bank_journal_xlsx
        wb = Workbook(); ws = wb.active
        ws.append(["账户：基本户"])
        ws.append(["日期", "摘要", "对方单位", "收入", "支出"])
        ws.append(["2026-06-05", "收货款", "客户甲", 500, None])
        ws.append([None, None, None, None, None])
        ws.append(["2026/06/06", "付款", "供应商甲", None, 200])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        parsed, errors = parse_bank_journal_xlsx(buf)
        self.assertEqual(len(parsed), 2)
        self.assertEqual(parsed[0]["direction"], "in")
        self.assertEqual(parsed[1]["direction"], "out")
        self.assertEqual(parsed[1]["date"], date(2026, 6, 6))


class BankJournalImportViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        U = get_user_model()
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户")
        cls.user = U.objects.create_user(username="fin", password="x", can_view_all_companies=True)
        for code in ("add_bankjournal", "view_bankjournal"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def _xlsx(self):
        from openpyxl import Workbook
        from io import BytesIO
        wb = Workbook(); ws = wb.active
        ws.append(["日期", "摘要", "对方单位", "收入", "支出"])
        ws.append(["2026-06-05", "收货款", "客户甲", 500, None])
        ws.append(["2026-06-06", "付电费", "电力公司", None, 80])
        buf = BytesIO(); wb.save(buf); return buf.getvalue()

    def test_import_creates_then_dedups(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from apps.finance.models import BankJournal
        self.client.force_login(self.user)
        data = self._xlsx()

        def upload():
            return self.client.post(
                "/finance/reports/bank-journal/import/",
                {"account": self.acc.pk,
                 "file": SimpleUploadedFile("流水.xlsx", data,
                     content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                SERVER_NAME="localhost", follow=True,
            )

        upload()
        self.assertEqual(BankJournal.objects.filter(company=self.c1, is_imported=True).count(), 2)
        upload()  # 第二次全部重复
        self.assertEqual(BankJournal.objects.filter(company=self.c1, is_imported=True).count(), 2)

    def _xlsx_with_txn(self, rows):
        """rows: [(date, summary, party, income, outcome, txn_no)]。"""
        from openpyxl import Workbook
        from io import BytesIO
        wb = Workbook(); ws = wb.active
        ws.append(["日期", "摘要", "对方单位", "收入", "支出", "交易流水号"])
        for r in rows:
            ws.append(list(r))
        buf = BytesIO(); wb.save(buf); return buf.getvalue()

    def _upload(self, data):
        from django.core.files.uploadedfile import SimpleUploadedFile
        return self.client.post(
            "/finance/reports/bank-journal/import/",
            {"account": self.acc.pk,
             "file": SimpleUploadedFile("流水.xlsx", data,
                 content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            SERVER_NAME="localhost", follow=True)

    def test_dedup_by_txn_no_even_if_content_changes(self):
        """有流水号时，按「账户+流水号」判重——即使摘要/金额变化也视为同一笔（修正重传）。"""
        from apps.finance.models import BankJournal
        self.client.force_login(self.user)
        self._upload(self._xlsx_with_txn([
            ("2026-06-05", "收货款", "客户甲", 500, None, "SN001"),
            ("2026-06-06", "付电费", "电力公司", None, 80, "SN002"),
        ]))
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 2)
        # 重传：SN001 摘要修正、新增 SN003；SN001/SN002 应跳过，仅新增 SN003
        self._upload(self._xlsx_with_txn([
            ("2026-06-05", "收货款(更正)", "客户甲", 500, None, "SN001"),
            ("2026-06-06", "付电费", "电力公司", None, 80, "SN002"),
            ("2026-06-07", "收利息", "银行", 3, None, "SN003"),
        ]))
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 3)
        self.assertEqual(BankJournal.objects.get(txn_no="SN001").summary, "收货款")  # 未被覆盖

    def test_dedup_within_same_batch(self):
        """同一文件内重复流水号只入一条。"""
        from apps.finance.models import BankJournal
        self.client.force_login(self.user)
        self._upload(self._xlsx_with_txn([
            ("2026-06-05", "收货款", "客户甲", 500, None, "SN001"),
            ("2026-06-05", "收货款", "客户甲", 500, None, "SN001"),
        ]))
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 1)


class PartnerlessReceiptPaymentTests(TestCase):
    """收/付款可不选往来对象（其他收款/其他付款）。"""

    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户",
                                             opening_balance=Decimal("1000"))

    def test_receipt_without_customer(self):
        from apps.finance.models import BankJournal
        from apps.finance.services import create_receipt
        rec = create_receipt(company=self.c1, user=None, doc_date=date(2026, 6, 6),
                             bank_account=self.acc, customer=None, amount=Decimal("200"),
                             summary="利息收入")
        self.assertIsNone(rec.customer)
        j = rec.bank_journal
        self.assertEqual(j.entry_type, BankJournal.EntryType.OTHER)  # 无往来 → 其他
        self.assertEqual(j.counterparty, "")
        self.assertEqual(j.amount, Decimal("200.00"))

    def test_payment_without_supplier(self):
        from apps.finance.models import BankJournal
        from apps.finance.services import create_payment
        pay = create_payment(company=self.c1, user=None, doc_date=date(2026, 6, 6),
                             bank_account=self.acc, supplier=None, amount=Decimal("50"),
                             summary="银行手续费")
        self.assertIsNone(pay.supplier)
        self.assertEqual(pay.bank_journal.entry_type, BankJournal.EntryType.OTHER)

    def test_receipt_with_customer_still_settlement(self):
        from apps.finance.models import BankJournal
        from apps.finance.services import create_receipt
        from apps.masterdata.models import Customer
        cust = Customer.objects.create(company=self.c1, code="C9", name="客户甲")
        rec = create_receipt(company=self.c1, user=None, doc_date=date(2026, 6, 6),
                             bank_account=self.acc, customer=cust, amount=Decimal("300"))
        self.assertEqual(rec.bank_journal.entry_type, BankJournal.EntryType.SETTLEMENT)


class OtherCashflowTests(TestCase):
    """其他收支登记（M8-2）：非往来银行收支直接成日记账。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户",
                                             opening_balance=Decimal("1000"))
        U = get_user_model()
        cls.user = U.objects.create_user(username="fin", password="x", can_view_all_companies=True)
        for code in ("add_bankjournal", "view_bankjournal", "delete_bankjournal"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def test_create_other_cashflow_makes_journal(self):
        from apps.finance.services import create_other_cashflow
        from apps.finance.models import BankJournal
        j = create_other_cashflow(
            company=self.c1, user=self.user, doc_date=date(2026, 6, 5),
            bank_account=self.acc, direction=BankJournal.Direction.OUT,
            amount=Decimal("88.50"), entry_type=BankJournal.EntryType.EXPENSE,
            counterparty="电力公司", summary="电费", txn_no="SN9")
        self.assertEqual(j.entry_type, "expense")
        self.assertEqual(j.source_type, "Other")
        self.assertEqual(j.amount, Decimal("88.50"))

    def test_settlement_type_rejected(self):
        from apps.finance.services import create_other_cashflow, SettlementError
        from apps.finance.models import BankJournal
        with self.assertRaises(SettlementError):
            create_other_cashflow(
                company=self.c1, user=self.user, doc_date=date(2026, 6, 5),
                bank_account=self.acc, direction=BankJournal.Direction.IN,
                amount=Decimal("100"), entry_type=BankJournal.EntryType.SETTLEMENT)

    def test_create_view_and_delete(self):
        from apps.finance.models import BankJournal
        self.client.force_login(self.user)
        r = self.client.post("/finance/other-cashflow/new/", {
            "doc_date": "2026-06-05", "bank_account": self.acc.pk, "direction": "out",
            "entry_type": "tax", "amount": "200", "counterparty": "税务局",
            "summary": "增值税", "txn_no": "",
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)
        j = BankJournal.objects.get(company=self.c1, entry_type="tax")
        self.assertEqual(j.amount, Decimal("200.00"))
        # 删除
        r2 = self.client.post(f"/finance/other-cashflow/{j.pk}/delete/",
                              SERVER_NAME="localhost", follow=True)
        self.assertEqual(r2.status_code, 200)
        self.assertFalse(BankJournal.objects.filter(pk=j.pk).exists())

    def test_cannot_delete_settlement_journal(self):
        from apps.finance.services import delete_other_cashflow, SettlementError
        pay = create_payment(company=self.c1, user=self.user, doc_date=date(2026, 6, 5),
                             bank_account=self.acc, supplier=Supplier.objects.create(
                                 company=self.c1, code="S9", name="供应商X"),
                             amount=Decimal("300"))
        with self.assertRaises(SettlementError):
            delete_other_cashflow(journal=pay.bank_journal, user=self.user)


class BankReconcileTests(TestCase):
    """银行对账（M8-3）：网银流水与日记账勾对、标记已对账、列出差异。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户",
                                             opening_balance=Decimal("1000"))
        # 系统已登记：一笔付款 300（流水号 SN1），一笔其他收支税费 200（无流水号）
        from apps.finance.services import create_other_cashflow
        from apps.finance.models import BankJournal
        cls.pay = create_payment(company=cls.c1, user=None, doc_date=date(2026, 6, 5),
                                 bank_account=cls.acc, supplier=cls.sup, amount=Decimal("300"))
        cls.pay.bank_journal.txn_no = "SN1"; cls.pay.bank_journal.save(update_fields=["txn_no"])
        create_other_cashflow(company=cls.c1, user=None, doc_date=date(2026, 6, 6),
                              bank_account=cls.acc, direction=BankJournal.Direction.OUT,
                              amount=Decimal("200"), entry_type=BankJournal.EntryType.TAX)
        cls.user = get_user_model().objects.create_user(
            username="fin", password="x", can_view_all_companies=True)

    def _reconcile(self, lines):
        from apps.finance.services import reconcile_bank_journal
        return reconcile_bank_journal(company=self.c1, user=self.user, account=self.acc,
                                      parsed=lines, filename="t.xlsx")

    def test_match_by_txn_and_by_content_plus_bank_only(self):
        from apps.finance.models import BankJournal
        lines = [
            # 匹配付款（按流水号 SN1）
            {"date": date(2026, 6, 5), "summary": "付货款", "counterparty": "供应商甲",
             "direction": "out", "amount": Decimal("300"), "txn_no": "SN1"},
            # 匹配税费（按 日期+金额+方向，无流水号）
            {"date": date(2026, 6, 6), "summary": "交税", "counterparty": "税务局",
             "direction": "out", "amount": Decimal("200"), "txn_no": ""},
            # 仅网银有：利息收入 5
            {"date": date(2026, 6, 30), "summary": "利息", "counterparty": "银行",
             "direction": "in", "amount": Decimal("5"), "txn_no": "SN9"},
        ]
        r = self._reconcile(lines)
        self.assertEqual(r["batch"].matched_count, 2)
        self.assertEqual(r["batch"].bank_only_count, 1)
        self.assertEqual(r["batch"].system_only_count, 0)
        # 匹配的日记账被标记已对账
        self.assertTrue(BankJournal.objects.get(txn_no="SN1").reconciled)
        self.assertEqual(r["bank_only"][0]["amount"], Decimal("5"))

    def test_system_only_listed(self):
        # 网银报了付款(06-05)和一笔利息(06-07)；期间内的税费(06-06)未出现在网银 → 仅系统有
        r = self._reconcile([
            {"date": date(2026, 6, 5), "summary": "付", "counterparty": "",
             "direction": "out", "amount": Decimal("300"), "txn_no": "SN1"},
            {"date": date(2026, 6, 7), "summary": "利息", "counterparty": "银行",
             "direction": "in", "amount": Decimal("5"), "txn_no": "SN9"},
        ])
        self.assertEqual(r["batch"].matched_count, 1)
        self.assertEqual(r["batch"].bank_only_count, 1)
        self.assertEqual(r["batch"].system_only_count, 1)  # 税费 200(06-06)


class InvoiceVoidTests(TestCase):
    """发票作废：未核销可作废(从应收/应付剔除)，已核销不可。"""

    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供A")

    def test_void_unsettled_purchase_invoice(self):
        from apps.finance.services import create_purchase_invoice, void_purchase_invoice_doc
        from apps.opening.reports import payable_partners_balance
        inv = create_purchase_invoice(company=self.c1, user=None, doc_date=date(2026, 6, 5),
            supplier=self.sup, lines=[{"product": None, "description": "货",
                                       "amount_untaxed": Decimal("1000"), "tax_rate": Decimal("0.13")}])
        void_purchase_invoice_doc(inv, None)
        inv.refresh_from_db()
        self.assertEqual(inv.status, "void")
        # 作废后从应付剔除
        rows = payable_partners_balance(self.c1, date(2026, 6, 1), date(2026, 6, 30))
        self.assertEqual(rows, [])

    def test_cannot_void_settled(self):
        from apps.finance.services import (create_purchase_invoice, create_payment,
            allocate_payment, void_purchase_invoice_doc, SettlementError)
        from apps.finance.models import BankAccount
        acc = BankAccount.objects.create(company=self.c1, name="基本户")
        inv = create_purchase_invoice(company=self.c1, user=None, doc_date=date(2026, 6, 5),
            supplier=self.sup, lines=[{"product": None, "description": "货",
                                       "amount_untaxed": Decimal("1000"), "tax_rate": Decimal("0")}])
        pay = create_payment(company=self.c1, user=None, doc_date=date(2026, 6, 6),
                             bank_account=acc, supplier=self.sup, amount=Decimal("1000"))
        allocate_payment(payment=pay, allocations=[{"invoice": inv, "amount": Decimal("1000")}])
        inv.refresh_from_db()
        with self.assertRaises(SettlementError):
            void_purchase_invoice_doc(inv, None)


class SalesRevenueCostTests(TestCase):
    """销售收入成本计算表（按开票口径、按商品）。"""

    @classmethod
    def setUpTestData(cls):
        from datetime import date
        from apps.inventory.services import post_inbound
        from apps.sales.services import create_and_post_outbound
        from apps.finance.services import create_sales_invoice
        from apps.masterdata.models import Customer
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="货A")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        # 进货 100@10（成本10）
        post_inbound(cls.c1, cls.p, Decimal("100"), Decimal("10"), date=date(2026, 6, 1))
        # 出库 30，售价含税单价 15 不含税
        ob = create_and_post_outbound(company=cls.c1, user=None, doc_date=date(2026, 6, 10),
            customer=cls.cust, lines=[{"product": cls.p, "quantity": Decimal("30"),
                                       "sale_unit_price": Decimal("15"), "tax_rate": Decimal("0")}])
        obl = ob.lines.first()
        # 发票关联该出库行（开票日 6/12）
        create_sales_invoice(company=cls.c1, user=None, doc_date=date(2026, 6, 12), customer=cls.cust,
            lines=[{"product": cls.p, "description": "", "amount_untaxed": Decimal("450"),
                    "tax_rate": Decimal("0"), "source_outbound_line": obl}])

    def test_revenue_cost_by_product(self):
        from datetime import date
        from apps.opening.reports import sales_revenue_cost
        d = sales_revenue_cost(self.c1, date(2026, 6, 1), date(2026, 6, 30))
        r = d["rows"][0]
        self.assertEqual(r["qty"], Decimal("30.000"))
        self.assertEqual(r["revenue"], Decimal("450.00"))      # 30×15
        self.assertEqual(r["cost"], Decimal("300.00"))         # 30×10 移动加权
        self.assertEqual(r["profit"], Decimal("150.00"))
        self.assertEqual(d["est_count"], 0)
        self.assertEqual(d["gap_count"], 0)

    def test_independent_with_quantity_estimates_cost(self):
        # 提前开票：未关联出库，但填了商品+数量 → 按移动加权单价(=10)估算成本
        from datetime import date
        from apps.finance.services import create_sales_invoice
        from apps.opening.reports import sales_revenue_cost
        create_sales_invoice(company=self.c1, user=None, doc_date=date(2026, 6, 20), customer=self.cust,
            lines=[{"product": self.p, "description": "", "quantity": Decimal("5"),
                    "amount_untaxed": Decimal("80"), "tax_rate": Decimal("0")}])
        d = sales_revenue_cost(self.c1, date(2026, 6, 1), date(2026, 6, 30))
        r = d["rows"][0]
        # 关联那张 30×10=300 + 估算 5×10=50 = 350 成本；数量 35
        self.assertEqual(r["cost"], Decimal("350.00"))
        self.assertEqual(r["qty"], Decimal("35.000"))
        self.assertEqual(d["est_count"], 1)
        self.assertEqual(d["gap_count"], 0)

    def test_gap_when_no_quantity(self):
        from datetime import date
        from apps.finance.services import create_sales_invoice
        from apps.opening.reports import sales_revenue_cost
        create_sales_invoice(company=self.c1, user=None, doc_date=date(2026, 6, 15), customer=self.cust,
            lines=[{"product": self.p, "description": "", "amount_untaxed": Decimal("100"),
                    "tax_rate": Decimal("0")}])  # 无关联、无数量
        d = sales_revenue_cost(self.c1, date(2026, 6, 1), date(2026, 6, 30))
        self.assertEqual(d["gap_count"], 1)
        self.assertEqual(d["gap_amount"], Decimal("100.00"))


class PartnerDrilldownTests(TestCase):
    """往来两级下钻（M9-2/M9-3）：供应商应付余额表 + 往来明细账。"""

    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户",
                                             opening_balance=Decimal("100000"))
        # 6/10 采购发票 1130（应付增）；6/20 付款 500 核销（应付减）
        cls.inv = create_purchase_invoice(
            company=cls.c1, user=None, doc_date=date(2026, 6, 10), supplier=cls.sup,
            lines=[{"product": None, "description": "货", "amount_untaxed": Decimal("1000"),
                    "tax_rate": Decimal("0.13")}])
        pay = create_payment(company=cls.c1, user=None, doc_date=date(2026, 6, 20),
                             bank_account=cls.acc, supplier=cls.sup, amount=Decimal("500"))
        allocate_payment(payment=pay, allocations=[{"invoice": cls.inv, "amount": Decimal("500")}])

    def test_payable_partners_balance(self):
        from apps.opening.reports import payable_partners_balance
        rows = payable_partners_balance(self.c1, date(2026, 6, 1), date(2026, 6, 30))
        r = next(x for x in rows if x["partner"] == self.sup)
        self.assertEqual(r["opening"], Decimal("0.00"))
        self.assertEqual(r["income"], Decimal("1130.00"))
        self.assertEqual(r["outgo"], Decimal("500.00"))
        self.assertEqual(r["ending"], Decimal("630.00"))

    def test_partner_ledger_rolling_balance(self):
        from apps.opening.reports import partner_ledger
        d = partner_ledger(self.c1, self.sup, "payable", date(2026, 6, 1), date(2026, 6, 30))
        self.assertEqual(d["opening"], Decimal("0.00"))
        self.assertEqual(d["income"], Decimal("1130.00"))
        self.assertEqual(d["outgo"], Decimal("500.00"))
        self.assertEqual(d["ending"], Decimal("630.00"))
        self.assertEqual(len(d["rows"]), 2)            # 发票 + 核销
        self.assertEqual(d["rows"][-1]["balance"], Decimal("630.00"))

    def test_ledger_period_before_shows_opening_only(self):
        from apps.opening.reports import partner_ledger
        # 区间在所有业务之后 → 全进期初，本期无发生，期初=期末
        d = partner_ledger(self.c1, self.sup, "payable", date(2026, 7, 1), date(2026, 7, 31))
        self.assertEqual(d["rows"], [])
        self.assertEqual(d["opening"], Decimal("630.00"))
        self.assertEqual(d["ending"], Decimal("630.00"))


class NoteDrilldownTests(TestCase):
    """应收票据两级下钻（M9-4）：票据余额表 + 使用明细。"""

    @classmethod
    def setUpTestData(cls):
        from apps.finance.services import (
            create_note_receivable, create_purchase_invoice, endorse_receivable_against_purchase,
        )
        from apps.masterdata.models import Customer, Supplier
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.cust = Customer.objects.create(company=cls.c1, code="C9", name="客户甲")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S9", name="供应商甲")
        # 6/5 出票 1000；用 600 背书抵一张采购发票（票出去→消耗未用额）
        cls.note = create_note_receivable(company=cls.c1, user=None, draw_date=date(2026, 6, 5),
                                          amount=Decimal("1000"), customer=cls.cust, note_no="BD001")
        pi = create_purchase_invoice(company=cls.c1, user=None, doc_date=date(2026, 6, 18),
            supplier=cls.sup, lines=[{"product": None, "description": "货",
                                      "amount_untaxed": Decimal("600"), "tax_rate": Decimal("0")}])
        endorse_receivable_against_purchase(note=cls.note,
                                            allocations=[{"invoice": pi, "amount": Decimal("600")}])

    def test_notes_balance(self):
        from apps.opening.reports import receivable_notes_balance
        rows = receivable_notes_balance(self.c1, date(2026, 6, 1), date(2026, 6, 30))
        r = rows[0]
        self.assertEqual(r["opening"], Decimal("0.00"))
        self.assertEqual(r["income"], Decimal("1000.00"))  # 本期出票
        self.assertEqual(r["outgo"], Decimal("600.00"))    # 本期背书(票出去)
        self.assertEqual(r["ending"], Decimal("400.00"))   # 未用

    def test_note_ledger(self):
        from apps.opening.reports import note_ledger
        d = note_ledger(self.c1, self.note, date(2026, 6, 1), date(2026, 6, 30))
        self.assertEqual(len(d["rows"]), 2)               # 出票 + 背书抵应付
        self.assertEqual(d["ending"], Decimal("400.00"))
        self.assertEqual(d["rows"][-1]["balance"], Decimal("400.00"))


class BankAccountsReportTests(TestCase):
    """银行存款分户余额表（总览下钻第一层）。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.acc1 = BankAccount.objects.create(company=cls.c1, name="基本户",
                                              opening_balance=Decimal("1000"))
        cls.acc2 = BankAccount.objects.create(company=cls.c1, name="一般户",
                                              opening_balance=Decimal("500"))
        # 基本户：期间内付款 300（流出）
        create_payment(company=cls.c1, user=None, doc_date=date(2026, 6, 5),
                       bank_account=cls.acc1, supplier=cls.sup, amount=Decimal("300"), summary="付款")
        U = get_user_model()
        cls.user = U.objects.create_user(username="fin", password="x", can_view_all_companies=True)
        cls.user.user_permissions.add(
            Permission.objects.get(content_type__app_label="finance", codename="view_bankjournal"))

    def _rows(self, dfrom="2026-06-01", dto="2026-06-30"):
        self.client.force_login(self.user)
        r = self.client.get(f"/finance/reports/bank-accounts/?company={self.c1.pk}&from={dfrom}&to={dto}",
                            SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 200)
        return {row["account"].name: row for row in r.context["rows"]}

    def test_per_account_balances(self):
        rows = self._rows()
        self.assertEqual(rows["基本户"]["opening"], Decimal("1000.00"))
        self.assertEqual(rows["基本户"]["outgo"], Decimal("300.00"))
        self.assertEqual(rows["基本户"]["ending"], Decimal("700.00"))
        # 一般户当期无发生：期初=期末=500
        self.assertEqual(rows["一般户"]["income"], Decimal("0.00"))
        self.assertEqual(rows["一般户"]["ending"], Decimal("500.00"))

    def test_future_range_shows_ending_without_activity(self):
        rows = self._rows("2099-01-01", "2099-12-31")
        # 付款已在区间前 → 计入期初；当期无发生但仍显示期末
        self.assertEqual(rows["基本户"]["income"], Decimal("0.00"))
        self.assertEqual(rows["基本户"]["outgo"], Decimal("0.00"))
        self.assertEqual(rows["基本户"]["ending"], Decimal("700.00"))
        self.assertEqual(rows["基本户"]["opening"], rows["基本户"]["ending"])


class NoteRegistrationTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")

    def test_create_notes(self):
        from apps.finance.services import create_note_receivable, create_note_payable
        nr = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 5),
                                    amount=Decimal("5000"), customer=self.cust, note_no="BA001")
        self.assertEqual(nr.doc_no, "YSP-C1-20260605-001")
        self.assertEqual(nr.unused, Decimal("5000.00"))
        self.assertEqual(nr.status, "on_hand")
        npay = create_note_payable(company=self.c1, user=None, draw_date=date(2026, 6, 5),
                                   supplier=self.sup, amount=Decimal("3000"))
        self.assertEqual(npay.doc_no, "YFP-C1-20260605-001")
        self.assertEqual(npay.unused, Decimal("3000.00"))


class NoteReceivableEditTests(TestCase):
    """应收票据修改/补录（到期日、来源客户、票号等）。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.p = Product.objects.create(company=cls.c1, code="P1", name="货A")
        U = get_user_model()
        cls.user = U.objects.create_user(username="cashier", password="x",
                                         can_view_all_companies=True)
        for code in ("add_notereceivable", "view_notereceivable"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def test_edit_supplements_missing_fields(self):
        """补录到期日/来源客户：导入或漏填的票据可补全。"""
        from apps.finance.services import create_note_receivable, update_note_receivable
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=Decimal("5000"))  # 无客户/无到期日/无票号
        self.assertIsNone(note.customer)
        update_note_receivable(note=note, user=self.user, draw_date=date(2026, 6, 11),
                               amount=Decimal("5000"), customer=self.cust,
                               note_no="BA999", due_date=date(2026, 9, 11), remark="补录")
        note.refresh_from_db()
        self.assertEqual(note.customer, self.cust)
        self.assertEqual(note.note_no, "BA999")
        self.assertEqual(note.due_date, date(2026, 9, 11))
        self.assertEqual(note.amount, Decimal("5000.00"))  # 未用，票面可保持

    def test_amount_locked_when_used(self):
        """已使用的票据（有冲销记录）票面金额锁定，但描述字段仍可补录。"""
        from apps.finance.services import (
            SettlementError, create_note_receivable, create_sales_invoice,
            settle_receivable_against_sales, update_note_receivable,
        )
        inv = create_sales_invoice(company=self.c1, user=None, doc_date=date(2026, 6, 11),
            customer=self.cust, lines=[{"product": self.p, "description": "",
            "amount_untaxed": Decimal("1000"), "tax_rate": Decimal("0")}])
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=Decimal("2000"), customer=self.cust)
        # 冲应收：产生使用记录（虽不消耗票面，但已与发票勾稽）→ 票面应锁定
        settle_receivable_against_sales(note=note,
            allocations=[{"invoice": inv, "amount": Decimal("1000")}])
        note.refresh_from_db()
        self.assertEqual(note.settled_amount, Decimal("0.00"))     # 核销应收不消耗票面
        # 改票面 → 拒绝（有冲销勾稽）
        with self.assertRaises(SettlementError):
            update_note_receivable(note=note, user=self.user, draw_date=note.draw_date,
                                   amount=Decimal("3000"), customer=self.cust)
        # 票面不变、只补到期日 → 通过
        update_note_receivable(note=note, user=self.user, draw_date=note.draw_date,
                               amount=Decimal("2000"), customer=self.cust,
                               due_date=date(2026, 12, 11))
        note.refresh_from_db()
        self.assertEqual(note.due_date, date(2026, 12, 11))

    def test_void_note_not_editable(self):
        from apps.finance.models import NoteReceivable
        from apps.finance.services import (
            SettlementError, create_note_receivable, update_note_receivable,
        )
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=Decimal("5000"), customer=self.cust)
        note.status = NoteReceivable.Status.VOID
        note.save(update_fields=["status"])
        with self.assertRaises(SettlementError):
            update_note_receivable(note=note, user=self.user, draw_date=note.draw_date,
                                   amount=Decimal("5000"))

    def test_edit_view_shows_button_and_saves(self):
        from apps.finance.services import create_note_receivable
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=Decimal("5000"))
        self.client.force_login(self.user)
        # 列表有「修改」按钮
        lst = self.client.get("/finance/notes-receivable/", SERVER_NAME="localhost")
        self.assertContains(lst, f"/finance/notes-receivable/{note.pk}/edit/")
        # GET 表单 200
        resp = self.client.get(f"/finance/notes-receivable/{note.pk}/edit/",
                               SERVER_NAME="localhost")
        self.assertEqual(resp.status_code, 200)
        # POST 补录
        resp = self.client.post(f"/finance/notes-receivable/{note.pk}/edit/", {
            "note_no": "BJ777", "draw_date": "2026-06-11", "due_date": "2026-09-11",
            "customer": self.cust.pk, "amount": "5000", "remark": "补录",
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(resp.status_code, 200)
        note.refresh_from_db()
        self.assertEqual(note.note_no, "BJ777")
        self.assertEqual(note.customer, self.cust)

    def test_used_amount_links_to_usage_detail(self):
        """已用金额（背书）可点 → 票据使用明细(all=1 看全部)，显示背书及被抵发票号。"""
        from apps.finance.services import (
            create_note_receivable, create_purchase_invoice, endorse_receivable_against_purchase,
        )
        pi = create_purchase_invoice(company=self.c1, user=None, doc_date=date(2026, 6, 11),
            supplier=self.sup, lines=[{"product": None, "description": "x",
            "amount_untaxed": Decimal("1000"), "tax_rate": Decimal("0")}])
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=Decimal("1000"), customer=self.cust)
        # 背书抵应付才消耗票面 → 已用>0
        endorse_receivable_against_purchase(note=note,
            allocations=[{"invoice": pi, "amount": Decimal("1000")}])
        self.client.force_login(self.user)
        # 列表「已用」是指向使用明细的链接
        lst = self.client.get("/finance/notes-receivable/", SERVER_NAME="localhost")
        self.assertContains(lst, "receivable-note-ledger")
        self.assertContains(lst, f"note={note.pk}")
        # 使用明细 all=1 显示背书事件 + 被抵发票号
        led = self.client.get(
            f"/finance/reports/receivable-note-ledger/?company={self.c1.pk}&note={note.pk}&all=1",
            SERVER_NAME="localhost")
        self.assertEqual(led.status_code, 200)
        self.assertContains(led, pi.doc_no)
        self.assertContains(led, "冲应收")


class NoteReceivableDeleteTests(TestCase):
    """应收票据删除（录错可彻底移除）：未使用、非期初才可删。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.p = Product.objects.create(company=cls.c1, code="P1", name="货A")
        U = get_user_model()
        cls.user = U.objects.create_user(username="cashier", password="x",
                                         can_view_all_companies=True)
        for code in ("add_notereceivable", "view_notereceivable", "view_receipt"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def test_delete_unused_note(self):
        from apps.finance.models import NoteReceivable
        from apps.finance.services import create_note_receivable, delete_note_receivable
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=Decimal("5000"), customer=self.cust)
        delete_note_receivable(note, user=self.user)
        self.assertFalse(NoteReceivable.objects.filter(pk=note.pk).exists())

    def test_used_note_blocked(self):
        """已使用（冲过应收）不可删——避免留下孤儿冲销记录。"""
        from apps.finance.services import (
            SettlementError, create_note_receivable, create_sales_invoice,
            delete_note_receivable, note_receivable_delete_block_reason,
            settle_receivable_against_sales,
        )
        inv = create_sales_invoice(company=self.c1, user=None, doc_date=date(2026, 6, 11),
            customer=self.cust, lines=[{"product": self.p, "description": "",
            "amount_untaxed": Decimal("1000"), "tax_rate": Decimal("0")}])
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=Decimal("2000"), customer=self.cust)
        settle_receivable_against_sales(note=note,
            allocations=[{"invoice": inv, "amount": Decimal("1000")}])
        note.refresh_from_db()
        self.assertIsNotNone(note_receivable_delete_block_reason(note))
        with self.assertRaises(SettlementError):
            delete_note_receivable(note, user=self.user)

    def test_opening_unused_note_deletable(self):
        """期初票据（导入最易录错）只要未使用即可删，不再额外拦期初。"""
        from apps.finance.models import NoteReceivable
        from apps.finance.services import create_note_receivable, delete_note_receivable
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 1),
                                      amount=Decimal("3000"), customer=self.cust, is_opening=True)
        delete_note_receivable(note, user=self.user)
        self.assertFalse(NoteReceivable.objects.filter(pk=note.pk).exists())

    def test_receipt_list_shows_delete_for_unused_note(self):
        """收款统一一览：未使用票据行出现「删除」表单，已使用的不出现。"""
        from apps.finance.services import (
            create_note_receivable, create_sales_invoice, settle_receivable_against_sales,
        )
        cls_p = self.p
        unused = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                        amount=Decimal("5000"), customer=self.cust)
        inv = create_sales_invoice(company=self.c1, user=None, doc_date=date(2026, 6, 11),
            customer=self.cust, lines=[{"product": cls_p, "description": "",
            "amount_untaxed": Decimal("1000"), "tax_rate": Decimal("0")}])
        used = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=Decimal("2000"), customer=self.cust)
        settle_receivable_against_sales(note=used,
            allocations=[{"invoice": inv, "amount": Decimal("1000")}])
        self.client.force_login(self.user)
        resp = self.client.get("/finance/receipts/", SERVER_NAME="localhost")
        self.assertContains(resp, f"/finance/notes-receivable/{unused.pk}/delete/")
        self.assertNotContains(resp, f"/finance/notes-receivable/{used.pk}/delete/")

    def test_delete_view_button_and_post(self):
        from apps.finance.models import NoteReceivable
        from apps.finance.services import create_note_receivable
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=Decimal("5000"))
        self.client.force_login(self.user)
        lst = self.client.get("/finance/notes-receivable/", SERVER_NAME="localhost")
        self.assertContains(lst, f"/finance/notes-receivable/{note.pk}/delete/")
        # GET 不删除（require_POST）
        self.assertEqual(self.client.get(
            f"/finance/notes-receivable/{note.pk}/delete/", SERVER_NAME="localhost").status_code, 405)
        # POST 删除
        resp = self.client.post(f"/finance/notes-receivable/{note.pk}/delete/",
                                SERVER_NAME="localhost", follow=True)
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(NoteReceivable.objects.filter(pk=note.pk).exists())


class NoteSettlementTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.p = Product.objects.create(company=cls.c1, code="P1", name="货A")

    def _sales_inv(self, untaxed):
        return create_sales_invoice(company=self.c1, user=None, doc_date=date(2026, 6, 5),
            customer=self.cust, lines=[{"product": self.p, "description": "",
            "amount_untaxed": Decimal(untaxed), "tax_rate": Decimal("0.13")}])

    def _purchase_inv(self, untaxed):
        return create_purchase_invoice(company=self.c1, user=None, doc_date=date(2026, 6, 5),
            supplier=self.sup, lines=[{"product": self.p, "description": "",
            "amount_untaxed": Decimal(untaxed), "tax_rate": Decimal("0.13")}])

    def test_receivable_note_settles_sales(self):
        from apps.finance.services import create_note_receivable, settle_receivable_against_sales
        inv = self._sales_inv("1000")  # 含税 1130
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 5),
                                      amount=Decimal("1130"), customer=self.cust)
        settle_receivable_against_sales(note=note,
            allocations=[{"invoice": inv, "amount": Decimal("1130")}])
        inv.refresh_from_db(); note.refresh_from_db()
        self.assertEqual(inv.outstanding, Decimal("0.00"))   # 应收账款已核销
        # 核销应收=票收进来抵应收，不消耗票面：票仍持有、可继续背书/托收
        self.assertEqual(note.unused, Decimal("1130.00"))
        self.assertEqual(note.settled_amount, Decimal("0.00"))
        self.assertEqual(note.status, "on_hand")

    def test_receivable_note_endorse_to_payable(self):
        from apps.finance.services import create_note_receivable, endorse_receivable_against_purchase
        pinv = self._purchase_inv("2000")  # 含税 2260
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 5),
                                      amount=Decimal("2260"), customer=self.cust)
        endorse_receivable_against_purchase(note=note,
            allocations=[{"invoice": pinv, "amount": Decimal("2260")}])
        pinv.refresh_from_db(); note.refresh_from_db()
        self.assertEqual(pinv.outstanding, Decimal("0.00"))
        self.assertEqual(note.status, "endorsed")

    def test_payable_note_settles_purchase_partial(self):
        from apps.finance.services import create_note_payable, settle_payable_against_purchase
        pinv = self._purchase_inv("2000")  # 含税 2260
        note = create_note_payable(company=self.c1, user=None, draw_date=date(2026, 6, 5),
                                   supplier=self.sup, amount=Decimal("5000"))
        settle_payable_against_purchase(note=note,
            allocations=[{"invoice": pinv, "amount": Decimal("2260")}])
        pinv.refresh_from_db(); note.refresh_from_db()
        self.assertEqual(pinv.outstanding, Decimal("0.00"))
        self.assertEqual(note.unused, Decimal("2740.00"))  # 5000-2260

    def test_note_over_use_rejected(self):
        from apps.finance.services import create_note_receivable, settle_receivable_against_sales, SettlementError
        inv = self._sales_inv("1000")  # 1130
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 5),
                                      amount=Decimal("500"), customer=self.cust)
        with self.assertRaises(SettlementError):
            settle_receivable_against_sales(note=note,
                allocations=[{"invoice": inv, "amount": Decimal("600")}])  # 超票据可用
        inv.refresh_from_db(); note.refresh_from_db()
        self.assertEqual(note.settled_amount, Decimal("0.00"))
        self.assertEqual(inv.settled_amount, Decimal("0.00"))


class NoteExcelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")

    def test_export_then_parse_notes(self):
        from io import BytesIO
        from apps.finance.services import create_note_receivable
        from apps.finance.excel import export_notes, parse_notes_xlsx
        create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 5),
                               amount=Decimal("2260"), customer=self.cust, note_no="BA1")
        notes = list(NoteReceivable.objects.filter(company=self.c1).select_related("customer"))
        content = export_notes(notes, "来源客户")
        parsed, errors = parse_notes_xlsx(BytesIO(content))
        self.assertEqual(errors, [])
        self.assertEqual(len(parsed), 1)
        self.assertEqual(parsed[0]["note_no"], "BA1")
        self.assertEqual(parsed[0]["amount"], Decimal("2260"))
        self.assertEqual(parsed[0]["party_name"], "K1 客户甲")


class ReceiptPaymentByNoteTests(TestCase):
    """收款方式=应收票据（收票冲应收）/ 付款方式=应收票据（背书抵应付）。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        U = get_user_model()
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.user = U.objects.create_user(username="cashier", password="x",
                                         can_view_all_companies=True)
        for code in ("add_receipt", "add_payment", "view_notereceivable"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def _sales_invoice(self, amount):
        return create_sales_invoice(
            company=self.c1, user=self.user, doc_date=date(2026, 6, 5), customer=self.cust,
            lines=[{"product": None, "description": "货", "amount_untaxed": amount,
                    "tax_rate": Decimal("0")}])

    def _purchase_invoice(self, amount):
        return create_purchase_invoice(
            company=self.c1, user=self.user, doc_date=date(2026, 6, 5), supplier=self.sup,
            lines=[{"product": None, "description": "货", "amount_untaxed": amount,
                    "tax_rate": Decimal("0")}])

    def test_receipt_by_note_creates_note_and_offsets_ar(self):
        inv = self._sales_invoice(Decimal("1000"))
        self.assertEqual(inv.outstanding, Decimal("1000.00"))
        self.client.force_login(self.user)
        r = self.client.post("/finance/receipts/new/", {
            "doc_date": "2026-06-10", "method": "note", "customer": self.cust.pk,
            "note_no": "BJ001", "draw_date": "2026-06-10", "due_date": "2026-09-10",
            "amount": "1000", "summary": "收客户票据",
            f"alloc-{inv.pk}": "1000",
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)
        note = NoteReceivable.objects.get(company=self.c1, note_no="BJ001")
        self.assertEqual(note.amount, Decimal("1000.00"))
        # 收票抵应收=票收进来抵应收账款，票仍持有（在手、未用=票面），可背书/托收
        self.assertEqual(note.status, NoteReceivable.Status.ON_HAND)
        self.assertEqual(note.unused, Decimal("1000.00"))
        inv.refresh_from_db()
        self.assertEqual(inv.outstanding, Decimal("0.00"))           # 应收已冲平
        # 不生成银行日记账
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 0)

    def test_receipt_by_note_partial_offset_keeps_remainder_on_hand(self):
        inv = self._sales_invoice(Decimal("600"))
        self.client.force_login(self.user)
        self.client.post("/finance/receipts/new/", {
            "doc_date": "2026-06-10", "method": "note", "customer": self.cust.pk,
            "note_no": "BJ002", "draw_date": "2026-06-10", "due_date": "2026-09-10",
            "amount": "1000", f"alloc-{inv.pk}": "600",
        }, SERVER_NAME="localhost", follow=True)
        note = NoteReceivable.objects.get(company=self.c1, note_no="BJ002")
        self.assertEqual(note.status, NoteReceivable.Status.ON_HAND)
        # 核销应收不消耗票：1000 票面全额仍持有（应收侧只冲了 600）
        self.assertEqual(note.unused, Decimal("1000.00"))
        inv.refresh_from_db()
        self.assertEqual(inv.outstanding, Decimal("0.00"))           # 应收 600 已冲平

    def test_payment_by_note_endorses_to_purchase(self):
        # 先有一张在手应收票据，再用它背书抵采购应付
        note = NoteReceivable.objects.create(
            company=self.c1, doc_no="YSP-C1-20260601-001", note_no="BJ010",
            draw_date=date(2026, 6, 1), due_date=date(2026, 9, 1),
            customer=self.cust, amount=Decimal("1000"))
        inv = self._purchase_invoice(Decimal("800"))
        self.client.force_login(self.user)
        r = self.client.post("/finance/payments/new/", {
            "doc_date": "2026-06-10", "method": "note", "supplier": self.sup.pk,
            "note_no": "BJ010", "amount": "800",
            f"alloc-{inv.pk}": "800",
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)
        note.refresh_from_db()
        # 部分背书：票面 1000 用了 800，余 200 仍「在手」可继续使用
        self.assertEqual(note.status, NoteReceivable.Status.ON_HAND)
        self.assertEqual(note.settled_amount, Decimal("800.00"))
        self.assertEqual(note.unused, Decimal("200.00"))
        inv.refresh_from_db()
        self.assertEqual(inv.outstanding, Decimal("0.00"))           # 应付已冲平
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 0)

    def test_payment_by_note_amount_must_match_allocations(self):
        NoteReceivable.objects.create(
            company=self.c1, doc_no="YSP-C1-20260601-002", note_no="BJ011",
            draw_date=date(2026, 6, 1), customer=self.cust, amount=Decimal("1000"))
        inv = self._purchase_invoice(Decimal("800"))
        self.client.force_login(self.user)
        self.client.post("/finance/payments/new/", {
            "doc_date": "2026-06-10", "method": "note", "supplier": self.sup.pk,
            "note_no": "BJ011", "amount": "900",   # 与勾选 800 不符 → 拒绝
            f"alloc-{inv.pk}": "800",
        }, SERVER_NAME="localhost", follow=True)
        inv.refresh_from_db()
        self.assertEqual(inv.outstanding, Decimal("800.00"))          # 未被冲销

    def test_edit_bank_payment_convert_to_note_endorsement(self):
        # 误记成银行付款 → 修改时切换为「应收票据(背书)」：删旧付款+日记账，改记背书
        from apps.finance.services import create_payment
        from django.utils import timezone
        acc = BankAccount.objects.create(company=self.c1, name="基本户")
        pay = create_payment(company=self.c1, user=self.user, doc_date=timezone.localdate(),
                            bank_account=acc, supplier=self.sup, amount=Decimal("800"),
                            summary="承兑")
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 1)
        note = NoteReceivable.objects.create(
            company=self.c1, doc_no="YSP-C1-20260601-009", note_no="BJ020",
            draw_date=date(2026, 6, 1), customer=self.cust, amount=Decimal("1000"))
        inv = self._purchase_invoice(Decimal("800"))
        self.client.force_login(self.user)
        r = self.client.post(f"/finance/payments/{pay.pk}/edit/", {
            "doc_date": timezone.localdate().strftime("%Y-%m-%d"),
            "method": "note", "supplier": self.sup.pk,
            "note_no": "BJ020", "amount": "800", f"alloc-{inv.pk}": "800",
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)
        # 原银行付款及其日记账已删除
        self.assertEqual(Payment.objects.filter(pk=pay.pk).count(), 0)
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 0)
        # 票据已背书抵应付
        note.refresh_from_db(); inv.refresh_from_db()
        self.assertEqual(note.settled_amount, Decimal("800.00"))
        self.assertEqual(inv.outstanding, Decimal("0.00"))

    def test_edit_convert_rolls_back_on_bad_allocation(self):
        # 切换票据时校验失败（合计≠付款金额）→ 原银行付款不应被删
        from apps.finance.services import create_payment
        from django.utils import timezone
        acc = BankAccount.objects.create(company=self.c1, name="基本户")
        pay = create_payment(company=self.c1, user=self.user, doc_date=timezone.localdate(),
                            bank_account=acc, supplier=self.sup, amount=Decimal("800"))
        NoteReceivable.objects.create(
            company=self.c1, doc_no="YSP-C1-20260601-010", note_no="BJ021",
            draw_date=date(2026, 6, 1), customer=self.cust, amount=Decimal("1000"))
        inv = self._purchase_invoice(Decimal("800"))
        self.client.force_login(self.user)
        self.client.post(f"/finance/payments/{pay.pk}/edit/", {
            "doc_date": timezone.localdate().strftime("%Y-%m-%d"),
            "method": "note", "supplier": self.sup.pk,
            "note_no": "BJ021", "amount": "900", f"alloc-{inv.pk}": "800",  # 不符
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(Payment.objects.filter(pk=pay.pk).count(), 1)       # 仍在
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 1)

    def test_edit_bank_receipt_convert_to_note(self):
        # 误记成银行收款 → 修改时切换为「应收票据」：删旧收款+日记账，改记为收到票据并冲应收
        acc = BankAccount.objects.create(company=self.c1, name="基本户")
        rec = create_receipt(company=self.c1, user=self.user, doc_date=date(2026, 6, 10),
                             bank_account=acc, customer=self.cust, amount=Decimal("1000"),
                             summary="承兑")
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 1)
        si = self._sales_invoice(Decimal("1000"))
        self.client.force_login(self.user)
        r = self.client.post(f"/finance/receipts/{rec.pk}/edit/", {
            "doc_date": "2026-06-10", "method": "note", "customer": self.cust.pk,
            "note_no": "BJ-CV1", "draw_date": "2026-06-10", "due_date": "2026-09-10",
            "amount": "1000", f"alloc-{si.pk}": "1000",
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)
        # 原银行收款及其日记账已删除
        self.assertEqual(Receipt.objects.filter(pk=rec.pk).count(), 0)
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 0)
        # 生成在手/已结算应收票据，且冲平应收
        note = NoteReceivable.objects.get(company=self.c1, note_no="BJ-CV1")
        self.assertEqual(note.amount, Decimal("1000.00"))
        si.refresh_from_db()
        self.assertEqual(si.outstanding, Decimal("0.00"))

    def test_bank_receipt_still_creates_journal(self):
        acc = BankAccount.objects.create(company=self.c1, name="基本户")
        self.client.force_login(self.user)
        self.client.post("/finance/receipts/new/", {
            "doc_date": "2026-06-10", "method": f"bank:{acc.pk}", "customer": "",
            "amount": "500", "summary": "现金收款",
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 1)
        self.assertEqual(NoteReceivable.objects.filter(company=self.c1).count(), 0)


class PurchaseInvoiceEditTests(TestCase):
    """采购发票修改：重算应付、已核销/期初/跨月拦截、视图入口。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        U = get_user_model()
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.sup2 = Supplier.objects.create(company=cls.c1, code="S2", name="供应商乙")
        cls.user = U.objects.create_user(username="purch", password="x",
                                         can_view_all_companies=True)
        for code in ("add_purchaseinvoice", "view_purchaseinvoice"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def _invoice(self, amount, day=None):
        from django.utils import timezone
        d = timezone.localdate() if day is None else day
        return create_purchase_invoice(
            company=self.c1, user=self.user, doc_date=d, supplier=self.sup,
            lines=[{"product": None, "description": "货", "amount_untaxed": amount,
                    "tax_rate": Decimal("0.13")}])

    def test_update_recomputes_payable(self):
        from apps.finance.services import update_purchase_invoice
        inv = self._invoice(Decimal("1000"))
        doc_no = inv.doc_no
        update_purchase_invoice(
            inv, user=self.user, doc_date=inv.doc_date, supplier=self.sup2,
            lines=[{"product": None, "description": "改后", "amount_untaxed": Decimal("2000"),
                    "tax_rate": Decimal("0.13")}])
        inv.refresh_from_db()
        self.assertEqual(inv.doc_no, doc_no)                 # 单号保留
        self.assertEqual(inv.supplier_id, self.sup2.pk)      # 供应商可改
        self.assertEqual(inv.amount_untaxed, Decimal("2000.00"))
        self.assertEqual(inv.tax_amount, Decimal("260.00"))
        self.assertEqual(inv.amount_taxed, Decimal("2260.00"))
        self.assertEqual(inv.lines.count(), 1)

    def test_settled_invoice_blocked(self):
        from apps.finance.services import update_purchase_invoice
        inv = self._invoice(Decimal("1000"))
        inv.settled_amount = Decimal("100")
        inv.save(update_fields=["settled_amount"])
        with self.assertRaises(SettlementError):
            update_purchase_invoice(
                inv, user=self.user, doc_date=inv.doc_date, supplier=self.sup,
                lines=[{"product": None, "description": "x", "amount_untaxed": Decimal("5"),
                        "tax_rate": Decimal("0")}])

    def test_block_reason_crossmonth_and_opening(self):
        from apps.finance.services import purchase_invoice_edit_block_reason
        today = date(2026, 6, 11)
        inv = self._invoice(Decimal("1000"), day=date(2026, 6, 8))
        self.assertIsNone(purchase_invoice_edit_block_reason(inv, today))   # 本月可改
        inv2 = self._invoice(Decimal("1000"), day=date(2026, 5, 8))
        self.assertEqual(purchase_invoice_edit_block_reason(inv2, today), "跨月发票不可修改")

    def test_edit_view_get_and_post(self):
        inv = self._invoice(Decimal("1000"))
        self.client.force_login(self.user)
        url = f"/finance/purchase-invoices/{inv.pk}/edit/"
        r = self.client.get(url, SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 200)
        r2 = self.client.post(url, {
            "doc_date": inv.doc_date.strftime("%Y-%m-%d"),
            "supplier": self.sup.pk, "invoice_no": "FP-NEW", "remark": "改备注", "term_days": "30",
            "form-TOTAL_FORMS": "1", "form-INITIAL_FORMS": "0",
            "form-MIN_NUM_FORMS": "0", "form-MAX_NUM_FORMS": "1000",
            "form-0-product": "", "form-0-description": "改后货", "form-0-quantity": "",
            "form-0-amount_untaxed": "1500", "form-0-tax_rate": "0.13",
            "form-0-tax_amount": "", "form-0-amount_taxed": "", "form-0-source_inbound_line": "",
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(r2.status_code, 200)
        inv.refresh_from_db()
        self.assertEqual(inv.invoice_no, "FP-NEW")
        self.assertEqual(inv.term_days, 30)
        self.assertEqual(inv.amount_untaxed, Decimal("1500.00"))
        self.assertEqual(inv.amount_taxed, Decimal("1695.00"))


class ReceiptPaymentEditDeleteTests(TestCase):
    """收款/付款 修改与删除：同步银行日记账、当月+未核销限制、删除即移除日记账。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        U = get_user_model()
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户")
        cls.acc2 = BankAccount.objects.create(company=cls.c1, name="一般户")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.user = U.objects.create_user(username="cash2", password="x",
                                         can_view_all_companies=True)
        for code in ("add_receipt", "add_payment", "view_receipt", "view_payment"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    # ---- service：修改同步日记账 ----
    def test_update_receipt_syncs_journal(self):
        from apps.finance.services import update_receipt
        from django.utils import timezone
        rec = create_receipt(company=self.c1, user=self.user, doc_date=timezone.localdate(),
                             bank_account=self.acc, customer=self.cust, amount=Decimal("500"))
        update_receipt(rec, user=self.user, doc_date=timezone.localdate(),
                       bank_account=self.acc2, customer=None, amount=Decimal("800"), summary="改")
        rec.refresh_from_db()
        self.assertEqual(rec.amount, Decimal("800.00"))
        self.assertEqual(rec.bank_account_id, self.acc2.pk)
        j = rec.bank_journal
        self.assertEqual(j.amount, Decimal("800.00"))           # 日记账同步
        self.assertEqual(j.bank_account_id, self.acc2.pk)
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 1)

    # ---- service：删除连带日记账 ----
    def test_delete_payment_removes_journal(self):
        from apps.finance.services import delete_payment
        from django.utils import timezone
        pay = create_payment(company=self.c1, user=self.user, doc_date=timezone.localdate(),
                            bank_account=self.acc, supplier=self.sup, amount=Decimal("600"))
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 1)
        delete_payment(pay, user=self.user)
        self.assertEqual(Payment.objects.filter(company=self.c1).count(), 0)
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 0)

    # ---- 已核销不可删 ----
    def test_settled_receipt_blocked(self):
        from apps.finance.services import delete_receipt, receipt_edit_block_reason
        from django.utils import timezone
        rec = create_receipt(company=self.c1, user=self.user, doc_date=timezone.localdate(),
                             bank_account=self.acc, customer=self.cust, amount=Decimal("500"))
        rec.settled_amount = Decimal("100"); rec.save(update_fields=["settled_amount"])
        self.assertIsNotNone(receipt_edit_block_reason(rec, timezone.localdate()))
        with self.assertRaises(SettlementError):
            delete_receipt(rec, user=self.user)

    # ---- 跨月不可改 ----
    def test_crossmonth_blocked(self):
        from apps.finance.services import receipt_edit_block_reason
        rec = create_receipt(company=self.c1, user=self.user, doc_date=date(2026, 5, 8),
                             bank_account=self.acc, customer=self.cust, amount=Decimal("500"))
        self.assertEqual(receipt_edit_block_reason(rec, date(2026, 6, 11)), "仅当月单据可修改/删除")

    # ---- 已对账不可改 ----
    def test_reconciled_blocked(self):
        from apps.finance.services import payment_edit_block_reason
        from django.utils import timezone
        pay = create_payment(company=self.c1, user=self.user, doc_date=timezone.localdate(),
                            bank_account=self.acc, supplier=self.sup, amount=Decimal("600"))
        j = pay.bank_journal; j.reconciled = True; j.save(update_fields=["reconciled"])
        pay.refresh_from_db()
        self.assertEqual(payment_edit_block_reason(pay, timezone.localdate()),
                         "该笔已银行对账，不可修改/删除")

    # ---- 视图：删除收款 ----
    def test_receipt_delete_view(self):
        from django.utils import timezone
        rec = create_receipt(company=self.c1, user=self.user, doc_date=timezone.localdate(),
                             bank_account=self.acc, customer=self.cust, amount=Decimal("500"))
        self.client.force_login(self.user)
        r = self.client.post(f"/finance/receipts/{rec.pk}/delete/",
                             SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)
        self.assertEqual(Receipt.objects.filter(company=self.c1).count(), 0)
        self.assertEqual(BankJournal.objects.filter(company=self.c1).count(), 0)

    # ---- 视图：修改付款 ----
    def test_payment_edit_view_post(self):
        from django.utils import timezone
        pay = create_payment(company=self.c1, user=self.user, doc_date=timezone.localdate(),
                            bank_account=self.acc, supplier=self.sup, amount=Decimal("600"))
        self.client.force_login(self.user)
        r = self.client.post(f"/finance/payments/{pay.pk}/edit/", {
            "doc_date": timezone.localdate().strftime("%Y-%m-%d"),
            "method": f"bank:{self.acc2.pk}", "supplier": self.sup.pk,
            "amount": "999", "summary": "改后",
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)
        pay.refresh_from_db()
        self.assertEqual(pay.amount, Decimal("999.00"))
        self.assertEqual(pay.bank_account_id, self.acc2.pk)
        self.assertEqual(pay.bank_journal.amount, Decimal("999.00"))


class OtherCashflowEditTests(TestCase):
    """其他收支：修改（仅 source_type=Other、未对账）。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        U = get_user_model()
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户")
        cls.user = U.objects.create_user(username="oc", password="x", can_view_all_companies=True)
        for code in ("add_bankjournal", "view_bankjournal"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def _other(self):
        from apps.finance.services import create_other_cashflow
        from django.utils import timezone
        return create_other_cashflow(
            company=self.c1, user=self.user, doc_date=timezone.localdate(),
            bank_account=self.acc, direction=BankJournal.Direction.OUT, amount=Decimal("200"),
            entry_type=BankJournal.EntryType.EXPENSE, summary="电费")

    def test_update_other_cashflow(self):
        from apps.finance.services import update_other_cashflow
        from django.utils import timezone
        j = self._other()
        update_other_cashflow(journal=j, user=self.user, doc_date=timezone.localdate(),
                              bank_account=self.acc, direction=BankJournal.Direction.OUT,
                              amount=Decimal("350"), entry_type=BankJournal.EntryType.EXPENSE,
                              summary="改电费")
        j.refresh_from_db()
        self.assertEqual(j.amount, Decimal("350.00"))
        self.assertEqual(j.summary, "改电费")

    def test_reconciled_other_cashflow_blocked(self):
        from apps.finance.services import update_other_cashflow
        from django.utils import timezone
        j = self._other()
        j.reconciled = True; j.save(update_fields=["reconciled"])
        with self.assertRaises(SettlementError):
            update_other_cashflow(journal=j, user=self.user, doc_date=timezone.localdate(),
                                  bank_account=self.acc, direction=BankJournal.Direction.OUT,
                                  amount=Decimal("350"), entry_type=BankJournal.EntryType.EXPENSE)

    def test_block_reason_crossmonth(self):
        from apps.finance.services import create_other_cashflow, other_cashflow_block_reason
        j = create_other_cashflow(
            company=self.c1, user=self.user, doc_date=date(2026, 5, 8),
            bank_account=self.acc, direction=BankJournal.Direction.OUT, amount=Decimal("200"),
            entry_type=BankJournal.EntryType.EXPENSE, summary="上月电费")
        self.assertEqual(other_cashflow_block_reason(j, date(2026, 6, 11)),
                         "仅当月单据可修改/删除")

    def test_non_other_journal_cannot_be_edited(self):
        # 往来生成的（如付款）source_type != Other，不可走此修改
        from apps.finance.services import create_payment, update_other_cashflow
        from apps.masterdata.models import Supplier
        from django.utils import timezone
        sup = Supplier.objects.create(company=self.c1, code="S1", name="供应商甲")
        pay = create_payment(company=self.c1, user=self.user, doc_date=timezone.localdate(),
                            bank_account=self.acc, supplier=sup, amount=Decimal("100"))
        with self.assertRaises(SettlementError):
            update_other_cashflow(journal=pay.bank_journal, user=self.user,
                                  doc_date=timezone.localdate(), bank_account=self.acc,
                                  direction=BankJournal.Direction.OUT, amount=Decimal("50"),
                                  entry_type=BankJournal.EntryType.EXPENSE)

    def test_edit_view_post(self):
        from django.utils import timezone
        j = self._other()
        self.client.force_login(self.user)
        r = self.client.post(f"/finance/other-cashflow/{j.pk}/edit/", {
            "doc_date": timezone.localdate().strftime("%Y-%m-%d"),
            "bank_account": self.acc.pk, "direction": "out",
            "entry_type": BankJournal.EntryType.EXPENSE, "amount": "500",
            "counterparty": "电力公司", "summary": "改后", "txn_no": "",
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)
        j.refresh_from_db()
        self.assertEqual(j.amount, Decimal("500.00"))
        self.assertEqual(j.counterparty, "电力公司")


class NoteMixedUseTests(TestCase):
    """一张应收票据可分次混合使用：部分冲应收 + 部分背书抵应付，未用完保持在手。"""

    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")

    def _note(self, amount):
        from apps.finance.services import create_note_receivable
        return create_note_receivable(
            company=self.c1, user=None, draw_date=date(2026, 6, 11),
            amount=Decimal(amount), customer=self.cust, note_no="BJ-MIX")

    def _sales(self, amount):
        return create_sales_invoice(
            company=self.c1, user=None, doc_date=date(2026, 6, 5), customer=self.cust,
            lines=[{"product": None, "description": "x", "amount_untaxed": Decimal(amount),
                    "tax_rate": Decimal("0")}])

    def _purchase(self, amount):
        return create_purchase_invoice(
            company=self.c1, user=None, doc_date=date(2026, 6, 5), supplier=self.sup,
            lines=[{"product": None, "description": "x", "amount_untaxed": Decimal(amount),
                    "tax_rate": Decimal("0")}])

    def test_settle_ar_does_not_consume_note_then_endorse(self):
        """收票抵应收账款不消耗票面（票留持有）；之后整张可背书出去。"""
        from apps.finance.services import (endorse_receivable_against_purchase,
                                           settle_receivable_against_sales)
        note = self._note("1000")
        si = self._sales("1000")
        pi = self._purchase("1000")

        # 冲应收 1000：减应收账款，但票不消耗（借应收票据/贷应收账款）
        settle_receivable_against_sales(
            note=note, allocations=[{"invoice": si, "amount": Decimal("1000")}])
        note.refresh_from_db(); si.refresh_from_db()
        self.assertEqual(note.status, NoteReceivable.Status.ON_HAND)   # 仍在手
        self.assertEqual(note.settled_amount, Decimal("0.00"))         # 未消耗
        self.assertEqual(note.unused, Decimal("1000.00"))             # 票面全额持有
        self.assertEqual(si.settled_amount, Decimal("1000.00"))       # 应收账款已冲

        # 持有的这张票再整额背书给供应商（票出去→消耗）
        endorse_receivable_against_purchase(
            note=note, allocations=[{"invoice": pi, "amount": Decimal("1000")}])
        note.refresh_from_db(); pi.refresh_from_db()
        self.assertEqual(note.unused, Decimal("0.00"))
        self.assertEqual(note.status, NoteReceivable.Status.ENDORSED) # 票出尽 → 已背书
        self.assertEqual(pi.settled_amount, Decimal("1000.00"))       # 应付已抵

    def test_partial_endorse_keeps_on_hand(self):
        from apps.finance.services import endorse_receivable_against_purchase
        note = self._note("1000")
        pi = self._purchase("1000")
        endorse_receivable_against_purchase(
            note=note, allocations=[{"invoice": pi, "amount": Decimal("300")}])
        note.refresh_from_db()
        self.assertEqual(note.status, NoteReceivable.Status.ON_HAND)   # 部分背书仍在手
        self.assertEqual(note.unused, Decimal("700.00"))

    def test_full_settle_ar_keeps_note_held(self):
        """纯冲应收全额：应收冲平，但票仍在手（不再变「已结算」）。"""
        from apps.finance.services import settle_receivable_against_sales
        note = self._note("500")
        si = self._sales("500")
        settle_receivable_against_sales(
            note=note, allocations=[{"invoice": si, "amount": Decimal("500")}])
        note.refresh_from_db()
        self.assertEqual(note.status, NoteReceivable.Status.ON_HAND)
        self.assertEqual(note.unused, Decimal("500.00"))
        self.assertEqual(note.settled_amount, Decimal("0.00"))

    def test_settle_ar_over_face_rejected(self):
        """冲应收合计不得超过票面（票面−已抵应收）。"""
        from apps.finance.services import SettlementError, settle_receivable_against_sales
        note = self._note("500")
        si = self._sales("1000")
        with self.assertRaises(SettlementError):
            settle_receivable_against_sales(
                note=note, allocations=[{"invoice": si, "amount": Decimal("600")}])


class InvoiceDeleteTests(TestCase):
    """采购/销售发票删除（彻底移除）：未核销、非期初才可删。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        U = get_user_model()
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.user = U.objects.create_user(username="inv", password="x", can_view_all_companies=True)
        for code in ("add_purchaseinvoice", "view_purchaseinvoice",
                     "add_salesinvoice", "view_salesinvoice"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def _purchase(self, amount=Decimal("1000")):
        from django.utils import timezone
        return create_purchase_invoice(
            company=self.c1, user=self.user, doc_date=timezone.localdate(), supplier=self.sup,
            lines=[{"product": None, "description": "x", "amount_untaxed": amount,
                    "tax_rate": Decimal("0.13")}])

    def test_delete_unsettled_purchase_invoice(self):
        from apps.finance.services import delete_purchase_invoice
        inv = self._purchase()
        delete_purchase_invoice(inv, user=self.user)
        self.assertEqual(PurchaseInvoice.objects.filter(company=self.c1).count(), 0)

    def test_settled_invoice_delete_blocked(self):
        from apps.finance.services import delete_purchase_invoice
        inv = self._purchase()
        inv.settled_amount = Decimal("100"); inv.save(update_fields=["settled_amount"])
        with self.assertRaises(SettlementError):
            delete_purchase_invoice(inv, user=self.user)
        self.assertEqual(PurchaseInvoice.objects.filter(company=self.c1).count(), 1)

    def test_opening_invoice_delete_blocked(self):
        from apps.finance.services import create_opening_payable, delete_purchase_invoice
        from django.utils import timezone
        inv = create_opening_payable(company=self.c1, user=self.user, supplier=self.sup,
                                     amount=Decimal("500"), doc_date=timezone.localdate())
        with self.assertRaises(SettlementError):
            delete_purchase_invoice(inv, user=self.user)

    def test_delete_voided_purchase_invoice_allowed(self):
        # 已作废且未核销的也可彻底删除（清理）
        from apps.finance.services import delete_purchase_invoice, void_purchase_invoice_doc
        inv = self._purchase()
        void_purchase_invoice_doc(inv, self.user)
        delete_purchase_invoice(inv, user=self.user)
        self.assertEqual(PurchaseInvoice.objects.filter(company=self.c1).count(), 0)

    def test_purchase_delete_view(self):
        inv = self._purchase()
        self.client.force_login(self.user)
        r = self.client.post(f"/finance/purchase-invoices/{inv.pk}/delete/",
                             SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)
        self.assertEqual(PurchaseInvoice.objects.filter(pk=inv.pk).count(), 0)

    def test_sales_delete_view(self):
        from django.utils import timezone
        inv = create_sales_invoice(
            company=self.c1, user=self.user, doc_date=timezone.localdate(), customer=self.cust,
            lines=[{"product": None, "description": "x", "amount_untaxed": Decimal("800"),
                    "tax_rate": Decimal("0")}])
        self.client.force_login(self.user)
        r = self.client.post(f"/finance/sales-invoices/{inv.pk}/delete/",
                             SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)
        self.assertEqual(SalesInvoice.objects.filter(pk=inv.pk).count(), 0)


class InvoiceSettlementBreakdownTests(TestCase):
    """发票详情「核销明细」：区分付款核销 vs 票据背书抵付。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        U = get_user_model()
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.user = U.objects.create_user(username="b", password="x", can_view_all_companies=True)
        for code in ("view_purchaseinvoice",):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def test_endorsement_shows_in_settlement_breakdown(self):
        from apps.finance.services import endorse_receivable_against_purchase
        from django.utils import timezone
        inv = create_purchase_invoice(
            company=self.c1, user=self.user, doc_date=timezone.localdate(), supplier=self.sup,
            lines=[{"product": None, "description": "x", "amount_untaxed": Decimal("1000"),
                    "tax_rate": Decimal("0")}])
        note = NoteReceivable.objects.create(
            company=self.c1, doc_no="YSP-C1-20260611-005", note_no="BJ232",
            draw_date=date(2026, 6, 11), customer=self.cust, amount=Decimal("232000.03"))
        endorse_receivable_against_purchase(
            note=note, allocations=[{"invoice": inv, "amount": Decimal("600")}], user=self.user)
        self.client.force_login(self.user)
        r = self.client.get(f"/finance/purchase-invoices/{inv.pk}/", SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 200)
        html = r.content.decode()
        self.assertIn("核销明细", html)
        self.assertIn("应收票据背书抵付", html)     # 核销方式标明是背书
        self.assertIn("YSP-C1-20260611-005", html)  # 指向那张票据


class UnifiedCashListTests(TestCase):
    """收款/付款列表统一一览：银行 + 应收票据都出现，方式列正确。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        U = get_user_model()
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="招商银行")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.user = U.objects.create_user(username="cl", password="x", can_view_all_companies=True)
        for code in ("view_receipt", "view_payment"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def test_receipt_list_includes_bank_and_note(self):
        from apps.finance.services import create_receipt
        from django.utils import timezone
        create_receipt(company=self.c1, user=self.user, doc_date=timezone.localdate(),
                       bank_account=self.acc, customer=self.cust, amount=Decimal("500"))
        NoteReceivable.objects.create(
            company=self.c1, doc_no="YSP-C1-20260611-005", note_no="BJ1",
            draw_date=date(2026, 6, 11), customer=self.cust, amount=Decimal("232000.03"))
        self.client.force_login(self.user)
        h = self.client.get("/finance/receipts/", SERVER_NAME="localhost").content.decode()
        self.assertIn("收款方式", h)
        self.assertIn("招商银行", h)                  # 银行收款行
        self.assertIn("应收票据", h)                  # 票据收款行
        self.assertIn("YSP-C1-20260611-005", h)
        self.assertIn("232,000.03", h)               # 千分位显示

    def test_payment_list_includes_bank_and_endorsement(self):
        from apps.finance.services import (create_payment, create_purchase_invoice,
                                           endorse_receivable_against_purchase)
        from django.utils import timezone
        create_payment(company=self.c1, user=self.user, doc_date=timezone.localdate(),
                       bank_account=self.acc, supplier=self.sup, amount=Decimal("600"))
        note = NoteReceivable.objects.create(
            company=self.c1, doc_no="YSP-C1-20260611-009", note_no="BJ9",
            draw_date=date(2026, 6, 11), customer=self.cust, amount=Decimal("1000"))
        inv = create_purchase_invoice(
            company=self.c1, user=self.user, doc_date=timezone.localdate(), supplier=self.sup,
            lines=[{"product": None, "description": "x", "amount_untaxed": Decimal("800"),
                    "tax_rate": Decimal("0")}])
        endorse_receivable_against_purchase(
            note=note, allocations=[{"invoice": inv, "amount": Decimal("800")}], user=self.user)
        self.client.force_login(self.user)
        h = self.client.get("/finance/payments/", SERVER_NAME="localhost").content.decode()
        self.assertIn("付款方式", h)
        self.assertIn("招商银行", h)                  # 银行付款行
        self.assertIn("应收票据背书", h)              # 背书付款行
        self.assertIn("YSP-C1-20260611-009", h)
        self.assertIn("供应商甲", h)                  # 背书行的供应商


class NoteSettlementReverseTests(TestCase):
    """撤销票据冲销：恢复发票未核销额 + 票据未用额/状态，救误用数据。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        U = get_user_model()
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.p = Product.objects.create(company=cls.c1, code="P1", name="货A")
        cls.user = U.objects.create_user(username="b", password="x", can_view_all_companies=True)
        for code in ("add_notesettlement", "view_salesinvoice", "view_purchaseinvoice",
                     "view_notereceivable"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def _settle_ar(self, note_amount, settle_amount):
        from apps.finance.services import create_note_receivable, create_sales_invoice, settle_receivable_against_sales
        inv = create_sales_invoice(company=self.c1, user=None, doc_date=date(2026, 6, 11),
            customer=self.cust, lines=[{"product": self.p, "description": "",
            "amount_untaxed": note_amount, "tax_rate": Decimal("0")}])
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=note_amount, customer=self.cust)
        settle_receivable_against_sales(note=note,
            allocations=[{"invoice": inv, "amount": settle_amount}])
        inv.refresh_from_db(); note.refresh_from_db()
        return inv, note

    def test_reverse_settle_ar_restores_invoice_only(self):
        """撤销「核销应收」：只退回发票未核销额；票本就没被消耗，未用额不变。"""
        from apps.finance.models import NoteReceivable, NoteSettlement
        from apps.finance.services import reverse_note_settlement
        inv, note = self._settle_ar(Decimal("1000"), Decimal("1000"))
        # 核销应收不消耗票：票在手、未用满、发票已核销
        self.assertEqual(note.status, NoteReceivable.Status.ON_HAND)
        self.assertEqual(note.unused, Decimal("1000.00"))
        self.assertEqual(inv.settled_amount, Decimal("1000.00"))
        s = NoteSettlement.objects.get(note_id=note.pk)
        reverse_note_settlement(settlement=s, user=self.user)
        inv.refresh_from_db(); note.refresh_from_db()
        # 撤销后：发票未核销额退回；票仍在手、未用不变（本就没消耗）
        self.assertEqual(note.status, NoteReceivable.Status.ON_HAND)
        self.assertEqual(note.unused, Decimal("1000.00"))
        self.assertEqual(inv.settled_amount, Decimal("0.00"))
        self.assertFalse(NoteSettlement.objects.filter(pk=s.pk).exists())

    def test_reverse_endorsement_restores_payable(self):
        from django.utils import timezone
        from apps.finance.models import NoteReceivable, NoteSettlement
        from apps.finance.services import (
            create_note_receivable, create_purchase_invoice,
            endorse_receivable_against_purchase, reverse_note_settlement,
        )
        pinv = create_purchase_invoice(company=self.c1, user=None, doc_date=timezone.localdate(),
            supplier=self.sup, lines=[{"product": None, "description": "x",
            "amount_untaxed": Decimal("800"), "tax_rate": Decimal("0")}])
        note = create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=Decimal("800"), customer=self.cust)
        endorse_receivable_against_purchase(
            note=note, allocations=[{"invoice": pinv, "amount": Decimal("800")}], user=self.user)
        pinv.refresh_from_db(); note.refresh_from_db()
        self.assertEqual(note.status, NoteReceivable.Status.ENDORSED)
        s = NoteSettlement.objects.get(note_id=note.pk, is_endorsement=True)
        reverse_note_settlement(settlement=s, user=self.user)
        pinv.refresh_from_db(); note.refresh_from_db()
        self.assertEqual(note.status, NoteReceivable.Status.ON_HAND)
        self.assertEqual(note.unused, Decimal("800.00"))
        self.assertEqual(pinv.settled_amount, Decimal("0.00"))

    def test_reverse_view_from_sales_invoice_detail(self):
        from apps.finance.models import NoteSettlement
        inv, note = self._settle_ar(Decimal("1000"), Decimal("1000"))
        s = NoteSettlement.objects.get(note_id=note.pk)
        self.client.force_login(self.user)
        detail = self.client.get(f"/finance/sales-invoices/{inv.pk}/", SERVER_NAME="localhost")
        self.assertEqual(detail.status_code, 200)
        self.assertContains(detail, f"/finance/note-settlements/{s.pk}/reverse/")
        # GET 不撤销（require_POST）
        self.assertEqual(self.client.get(
            f"/finance/note-settlements/{s.pk}/reverse/", SERVER_NAME="localhost").status_code, 405)
        # POST 撤销
        resp = self.client.post(f"/finance/note-settlements/{s.pk}/reverse/",
                                {"next": f"/finance/sales-invoices/{inv.pk}/"},
                                SERVER_NAME="localhost", follow=True)
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(NoteSettlement.objects.filter(pk=s.pk).exists())
        inv.refresh_from_db()
        self.assertEqual(inv.settled_amount, Decimal("0.00"))


class ListTotalsTests(TestCase):
    """应收票据 / 收款登记 / 付款登记 列表底部「合计」行。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        from apps.finance.models import BankAccount
        from apps.finance.services import (
            create_note_receivable, create_payment, create_receipt,
        )
        U = get_user_model()
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户")
        cls.user = U.objects.create_user(username="fin", password="x", can_view_all_companies=True)
        for code in ("view_notereceivable", "view_receipt", "view_payment"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))
        d = date(2026, 6, 11)
        create_note_receivable(company=cls.c1, user=None, draw_date=d,
                               amount=Decimal("1000"), customer=cls.cust)
        create_note_receivable(company=cls.c1, user=None, draw_date=d,
                               amount=Decimal("500"), customer=cls.cust)
        create_receipt(company=cls.c1, user=None, doc_date=d, bank_account=cls.acc,
                       customer=cls.cust, amount=Decimal("300"))
        create_receipt(company=cls.c1, user=None, doc_date=d, bank_account=cls.acc,
                       customer=cls.cust, amount=Decimal("200"))
        create_payment(company=cls.c1, user=None, doc_date=d, bank_account=cls.acc,
                       supplier=cls.sup, amount=Decimal("700"))

    def test_note_list_totals(self):
        from apps.finance.services import create_note_receivable
        self.client.force_login(self.user)
        resp = self.client.get("/finance/notes-receivable/", SERVER_NAME="localhost")
        self.assertEqual(resp.status_code, 200)
        t = resp.context["totals"]
        # 两张均本期出票（非期初）→ 期初金额 0、本期收入 1500
        self.assertEqual(t["opening"], Decimal("0.00"))
        self.assertEqual(t["period"], Decimal("1500.00"))
        self.assertEqual(t["unused"], Decimal("1500.00"))   # 均在手
        self.assertContains(resp, "期初金额")
        self.assertContains(resp, "本期收入")
        self.assertContains(resp, "合计")
        # 加一张期初票据 → 落入「期初金额」列合计
        create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 1),
                               amount=Decimal("800"), customer=self.cust, is_opening=True)
        t2 = self.client.get("/finance/notes-receivable/", SERVER_NAME="localhost").context["totals"]
        self.assertEqual(t2["opening"], Decimal("800.00"))
        self.assertEqual(t2["period"], Decimal("1500.00"))

    def test_receipt_list_totals(self):
        self.client.force_login(self.user)
        resp = self.client.get("/finance/receipts/", SERVER_NAME="localhost")
        self.assertEqual(resp.status_code, 200)
        # 统一一览：银行收款 300+200 + 收到应收票据 1000+500 = 2000
        self.assertEqual(resp.context["totals"]["amount"], Decimal("2000.00"))
        self.assertContains(resp, "合计")

    def test_payment_list_totals(self):
        self.client.force_login(self.user)
        resp = self.client.get("/finance/payments/", SERVER_NAME="localhost")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["totals"]["amount"], Decimal("700.00"))
        self.assertContains(resp, "合计")


class NoteCashTests(TestCase):
    """应收票据到期兑付 / 贴现：票据→银行存款（贴现多记财务费用）。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        from apps.finance.models import BankAccount
        U = get_user_model()
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.cust = Customer.objects.create(company=cls.c1, code="K1", name="客户甲")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户")
        cls.user = U.objects.create_user(username="fin", password="x", can_view_all_companies=True)
        for code in ("add_notesettlement", "view_notereceivable"):
            cls.user.user_permissions.add(
                Permission.objects.get(content_type__app_label="finance", codename=code))

    def _note(self, amount):
        from apps.finance.services import create_note_receivable
        return create_note_receivable(company=self.c1, user=None, draw_date=date(2026, 6, 11),
                                      amount=Decimal(amount), customer=self.cust)

    def test_collect_creates_bank_journal_consumes_note(self):
        from apps.finance.models import BankJournal, NoteReceivable
        from apps.finance.services import collect_note_receivable
        note = self._note("1000")
        collect_note_receivable(note=note, user=self.user, date=date(2026, 9, 11),
                                bank_account=self.acc, amount=Decimal("1000"))
        note.refresh_from_db()
        self.assertEqual(note.unused, Decimal("0.00"))           # 票变现金、消耗
        self.assertEqual(note.status, NoteReceivable.Status.SETTLED)
        j = BankJournal.objects.get(company=self.c1, source_type="NoteDisposal")
        self.assertEqual(j.direction, BankJournal.Direction.IN)
        self.assertEqual(j.amount, Decimal("1000.00"))           # 票面进银行
        self.assertEqual(j.entry_type, BankJournal.EntryType.NOTE_CASH)

    def test_discount_nets_cash_and_records_finance_expense(self):
        from apps.finance.models import BankJournal, ExpenseRecord
        from apps.finance.services import discount_note_receivable
        note = self._note("1000")
        discount_note_receivable(note=note, user=self.user, date=date(2026, 7, 1),
                                 bank_account=self.acc, net_amount=Decimal("992"),
                                 amount=Decimal("1000"))
        note.refresh_from_db()
        self.assertEqual(note.unused, Decimal("0.00"))
        j = BankJournal.objects.get(company=self.c1, source_type="NoteDisposal")
        self.assertEqual(j.amount, Decimal("992.00"))            # 实收净额进银行
        exp = ExpenseRecord.objects.get(company=self.c1, category=ExpenseRecord.Category.FINANCE)
        self.assertEqual(exp.amount, Decimal("8.00"))            # 贴现息=1000-992

    def test_reverse_disposal_restores_note_and_deletes_journal(self):
        from apps.finance.models import BankJournal, ExpenseRecord, NoteDisposal, NoteReceivable
        from apps.finance.services import discount_note_receivable, reverse_note_disposal
        note = self._note("1000")
        d = discount_note_receivable(note=note, user=self.user, date=date(2026, 7, 1),
                                     bank_account=self.acc, net_amount=Decimal("992"),
                                     amount=Decimal("1000"))
        reverse_note_disposal(disposal=d, user=self.user)
        note.refresh_from_db()
        self.assertEqual(note.unused, Decimal("1000.00"))        # 恢复持有
        self.assertEqual(note.status, NoteReceivable.Status.ON_HAND)
        self.assertFalse(NoteDisposal.objects.filter(pk=d.pk).exists())
        self.assertEqual(BankJournal.objects.filter(company=self.c1, source_type="NoteDisposal").count(), 0)
        self.assertEqual(ExpenseRecord.objects.filter(company=self.c1).count(), 0)

    def test_note_balance_counts_disposal_as_outgo(self):
        """票据余额表：兑付/贴现也算「本期发出」（票出去）。"""
        from apps.opening.reports import receivable_notes_balance
        from apps.finance.services import collect_note_receivable
        note = self._note("1000")
        collect_note_receivable(note=note, user=self.user, date=date(2026, 6, 20),
                                bank_account=self.acc, amount=Decimal("1000"))
        rows = receivable_notes_balance(self.c1, date(2026, 6, 1), date(2026, 6, 30))
        r = rows[0]
        self.assertEqual(r["income"], Decimal("1000.00"))        # 出票
        self.assertEqual(r["outgo"], Decimal("1000.00"))         # 兑付(票出去)
        self.assertEqual(r["ending"], Decimal("0.00"))

    def test_collect_view_button_and_post(self):
        from apps.finance.models import NoteDisposal
        note = self._note("500")
        self.client.force_login(self.user)
        lst = self.client.get("/finance/notes-receivable/", SERVER_NAME="localhost")
        self.assertContains(lst, f"/finance/notes-receivable/{note.pk}/collect/")
        self.assertContains(lst, f"/finance/notes-receivable/{note.pk}/discount/")
        resp = self.client.post(f"/finance/notes-receivable/{note.pk}/collect/", {
            "bank_account": self.acc.pk, "date": "2026-09-11", "amount": "500",
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(NoteDisposal.objects.filter(note=note).count(), 1)
