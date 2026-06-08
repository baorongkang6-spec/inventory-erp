"""银行存款日记账的 Excel 导出 / 导入（openpyxl）。SPEC §7.1。

导入格式（首行表头，列名可按需扩展）：
    日期 | 摘要 | 对方单位 | 收入 | 支出
- 日期接受 Excel 日期或 YYYY-MM-DD / YYYY/MM/DD 文本。
- 收入>0 记为收入方向；否则按支出>0 记为支出方向。
"""

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import Workbook, load_workbook

HEADERS = ["日期", "摘要", "对方单位", "收入", "支出", "交易流水号", "余额"]
# 导入读取前 6 列（余额为只读统计列，不参与导入）。
IMPORT_COLS = 6


# 导出时这些汇总行写在数据区上下；再导入时按首列识别并跳过（不当作流水）。
SUMMARY_LABELS = {"期初余额", "本期合计", "合计", "期末余额"}


def export_bank_journal(account, rows, opening=None, closing=None,
                        date_from=None, date_to=None) -> bytes:
    """rows: [{"j": BankJournal, "balance": Decimal}]。

    含 期初余额 / 逐笔 / 本期合计(收入、支出) / 期末余额 行，便于直接对账。
    opening/closing 缺省时按首末行余额回推。
    """
    from decimal import Decimal as D
    wb = Workbook()
    ws = wb.active
    ws.title = "银行存款日记账"

    period = ""
    if date_from or date_to:
        period = f"  期间：{date_from or '起初'} ~ {date_to or '至今'}"
    ws.append([account.company.header_name])   # 抬头：公司法定全称
    ws.append([f"账户：{account.name} {account.account_no}{period}"])
    ws.append(HEADERS)
    # 左上角嵌入 logo（缺 Pillow/文件不存在则跳过）
    try:
        from django.conf import settings
        from openpyxl.drawing.image import Image as XLImage
        lp = settings.BASE_DIR / "static" / "img" / "logo.png"
        if lp.exists():
            im = XLImage(str(lp)); im.height, im.width = 34, round(34 * im.width / im.height)
            ws.add_image(im, "A1")
    except Exception:
        pass

    if opening is None:
        opening = (rows[0]["balance"] - rows[0]["j"].signed_amount) if rows else account.opening_balance
    if closing is None:
        closing = rows[-1]["balance"] if rows else opening

    # 期初余额行（余额列）
    ws.append(["期初余额", None, None, None, None, None, float(opening)])

    income_total = outgo_total = D("0.00")
    for r in rows:
        j = r["j"]
        is_in = j.direction == "in"
        if is_in:
            income_total += j.amount
        else:
            outgo_total += j.amount
        ws.append([
            j.date.strftime("%Y-%m-%d"),
            j.summary,
            j.counterparty,
            float(j.amount) if is_in else None,
            float(j.amount) if not is_in else None,
            j.txn_no,
            float(r["balance"]),
        ])

    # 本期合计行（收入、支出合计）
    ws.append(["本期合计", None, None, float(income_total), float(outgo_total), None, None])
    # 期末余额行（余额列）
    ws.append(["期末余额", None, None, None, None, None, float(closing)])

    # 金额列（收入D / 支出E / 余额G）统一两位小数 + 千分位
    for row in ws.iter_rows(min_row=4, min_col=1, max_col=7):
        for cell in row:
            if cell.column_letter in ("D", "E", "G") and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_bank_template() -> bytes:
    """空白导入模板（含表头与一行示例）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "银行流水导入"
    ws.append(HEADERS[:IMPORT_COLS])
    ws.append(["2026-06-01", "示例：货款收入", "某某公司", 10000, None, "SN20260601001"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _to_date(value):
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    if value is None:
        return None
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_decimal(value):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def parse_bank_journal_xlsx(file):
    """解析上传的 xlsx，返回 (有效行列表, 错误信息列表)。

    每行 dict：{date, summary, counterparty, direction, amount}。
    跳过空行与金额为 0 的行；表头行（含「日期」）自动识别跳过。
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    parsed, errors = [], []
    for idx, values in enumerate(ws.iter_rows(values_only=True), start=1):
        values = list(values)
        if not any(v not in (None, "") for v in values):
            continue
        # 跳过标题/表头行与导出的汇总行（期初余额/本期合计/期末余额）
        first = str(values[0]).strip() if values and values[0] is not None else ""
        if first.startswith("账户") or first == "日期" or first in SUMMARY_LABELS:
            continue
        # 期望至少 6 列：日期 摘要 对方 收入 支出 交易流水号
        cells = (values + [None] * IMPORT_COLS)[:IMPORT_COLS]
        d = _to_date(cells[0])
        if d is None:
            if not parsed:   # 尚无数据行 → 视为前置抬头/标题行，静默跳过
                continue
            errors.append(f"第 {idx} 行：日期无法识别（{cells[0]!r}）")
            continue
        income = _to_decimal(cells[3])
        outcome = _to_decimal(cells[4])
        if income > 0:
            direction, amount = "in", income
        elif outcome > 0:
            direction, amount = "out", outcome
        else:
            errors.append(f"第 {idx} 行：收入/支出均为 0，已跳过")
            continue
        parsed.append({
            "date": d,
            "summary": str(cells[1] or "").strip(),
            "counterparty": str(cells[2] or "").strip(),
            "direction": direction,
            "amount": amount,
            "txn_no": str(cells[5] or "").strip(),
        })
    return parsed, errors


# 前 5 列与导入格式一致（票据号/出票日/到期日/对方/票面），其后为只读统计列，
# 便于「导出→编辑→再导入」往返；导入只读前 5 列。
NOTE_HEADERS = ["票据号", "出票日", "到期日", "对方单位", "票面金额", "已用", "未用", "状态", "单号"]
NOTE_IMPORT_HINT = ["票据号", "出票日", "到期日", "对方单位", "票面金额"]


def export_notes(notes, party_label="对方单位") -> bytes:
    """导出票据台账（应收/应付通用）。notes 为 NoteReceivable/NotePayable 列表。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "票据台账"
    headers = NOTE_HEADERS.copy()
    headers[3] = party_label
    ws.append(headers)
    for n in notes:
        party = getattr(n, "customer", None) or getattr(n, "supplier", None)
        ws.append([
            n.note_no,
            n.draw_date.strftime("%Y-%m-%d") if n.draw_date else "",
            n.due_date.strftime("%Y-%m-%d") if n.due_date else "",
            str(party) if party else "",
            float(n.amount), float(n.settled_amount), float(n.unused),
            n.get_status_display(), n.doc_no,
        ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_notes_xlsx(file):
    """解析票据导入 xlsx，列：票据号 | 出票日 | 到期日 | 对方单位 | 票面金额。

    返回 (有效行, 错误)。每行 dict：{note_no, draw_date, due_date, party_name, amount}。
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    parsed, errors = [], []
    for idx, values in enumerate(ws.iter_rows(values_only=True), start=1):
        values = list(values)
        if not any(v not in (None, "") for v in values):
            continue
        first = str(values[0]).strip() if values and values[0] is not None else ""
        if first in ("票据号", "单号") or first.startswith("账户"):
            continue
        cells = (values + [None] * 5)[:5]
        draw = _to_date(cells[1])
        if draw is None:
            errors.append(f"第 {idx} 行：出票日无法识别（{cells[1]!r}）")
            continue
        amount = _to_decimal(cells[4])
        if amount <= 0:
            errors.append(f"第 {idx} 行：票面金额无效，已跳过")
            continue
        parsed.append({
            "note_no": str(cells[0] or "").strip(),
            "draw_date": draw,
            "due_date": _to_date(cells[2]),
            "party_name": str(cells[3] or "").strip(),
            "amount": amount,
        })
    return parsed, errors
