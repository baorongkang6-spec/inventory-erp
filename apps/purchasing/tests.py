"""采购入库过账集成测试。"""

from datetime import date
from decimal import Decimal

from django.test import TestCase

from apps.core.models import Company
from apps.inventory.models import StockBalance
from apps.masterdata.models import Product
from apps.purchasing.models import PurchaseInbound
from apps.purchasing.services import create_and_post_inbound


class InboundPostingTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="环氧树脂")

    def test_create_and_post_weighted_average(self):
        doc = create_and_post_inbound(
            company=self.c1, user=None, doc_date=date(2026, 6, 5),
            lines=[
                {"product": self.p, "quantity": Decimal("100"), "unit_price": Decimal("10")},
                {"product": self.p, "quantity": Decimal("50"), "unit_price": Decimal("13")},
            ],
        )
        self.assertEqual(doc.doc_no, "RK-C1-20260605-001")
        self.assertEqual(doc.total_quantity, Decimal("150.000"))
        self.assertEqual(doc.total_amount, Decimal("1650.00"))
        self.assertEqual(doc.lines.count(), 2)

        bal = StockBalance.objects.get(company=self.c1, product=self.p)
        self.assertEqual(bal.quantity, Decimal("150.000"))
        self.assertEqual(bal.amount, Decimal("1650.00"))
        self.assertEqual(bal.avg_price, Decimal("11.00"))

    def test_doc_no_increments_per_day(self):
        for _ in range(2):
            create_and_post_inbound(
                company=self.c1, user=None, doc_date=date(2026, 6, 5),
                lines=[{"product": self.p, "quantity": Decimal("1"), "unit_price": Decimal("1")}],
            )
        nos = list(PurchaseInbound.objects.order_by("doc_no").values_list("doc_no", flat=True))
        self.assertEqual(nos, ["RK-C1-20260605-001", "RK-C1-20260605-002"])


class TaxInclusivePriceTests(TestCase):
    """含税单价录入：自动反算不含税/税额（M15）。"""

    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="货A")

    def test_inbound_from_tax_inclusive_price(self):
        from apps.purchasing.services import create_and_post_inbound
        doc = create_and_post_inbound(company=self.c1, user=None, doc_date=date(2026, 6, 5),
            lines=[{"product": self.p, "quantity": Decimal("10"),
                    "tax_inclusive_price": Decimal("113"), "tax_rate": Decimal("0.13")}])
        ln = doc.lines.first()
        self.assertEqual(ln.amount_taxed, Decimal("1130.00"))
        self.assertEqual(ln.amount_untaxed, Decimal("1000.00"))   # 1130/1.13
        self.assertEqual(ln.tax_amount, Decimal("130.00"))
        self.assertEqual(ln.amount, Decimal("1000.00"))           # 入库成本=不含税

    def test_explicit_amounts_win(self):
        from apps.purchasing.services import create_and_post_inbound
        # 手工改了金额（含税单价仅作参考）→ 用显式金额
        doc = create_and_post_inbound(company=self.c1, user=None, doc_date=date(2026, 6, 5),
            lines=[{"product": self.p, "quantity": Decimal("10"), "tax_rate": Decimal("0.13"),
                    "tax_inclusive_price": Decimal("113"),
                    "amount_untaxed": Decimal("990"), "tax_amount": Decimal("128.70"),
                    "amount_taxed": Decimal("1118.70")}])
        ln = doc.lines.first()
        self.assertEqual(ln.amount_untaxed, Decimal("990.00"))
        self.assertEqual(ln.amount_taxed, Decimal("1118.70"))


class InboundEditTests(TestCase):
    """采购入库修改（M14）：冲正重过账、保留单号、可改性守卫。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="货A")
        cls.u = get_user_model().objects.create_user(username="op", password="x")

    def _today(self):
        from django.utils import timezone
        return timezone.localdate()

    def test_update_reposts_and_keeps_docno(self):
        from apps.purchasing.services import create_and_post_inbound, update_and_repost_inbound
        doc = create_and_post_inbound(company=self.c1, user=self.u, doc_date=self._today(),
            lines=[{"product": self.p, "quantity": Decimal("10"), "unit_price": Decimal("5")}])
        no = doc.doc_no
        update_and_repost_inbound(doc, user=self.u, doc_date=self._today(),
            lines=[{"product": self.p, "quantity": Decimal("20"), "unit_price": Decimal("6"),
                    "tax_rate": Decimal("0.13")}])
        doc.refresh_from_db()
        self.assertEqual(doc.doc_no, no)                  # 单号不变
        self.assertEqual(doc.total_quantity, Decimal("20.000"))
        bal = StockBalance.objects.get(company=self.c1, product=self.p)
        self.assertEqual(bal.quantity, Decimal("20.000"))
        self.assertEqual(bal.amount, Decimal("120.00"))

    def test_block_reasons(self):
        from datetime import date
        from apps.purchasing.services import create_and_post_inbound, inbound_edit_block_reason
        doc = create_and_post_inbound(company=self.c1, user=self.u, doc_date=self._today(),
            lines=[{"product": self.p, "quantity": Decimal("1"), "unit_price": Decimal("1")}])
        from django.contrib.auth import get_user_model
        other = get_user_model().objects.create_user(username="x2", password="x")
        self.assertIsNone(inbound_edit_block_reason(doc, self.u, self._today(), False))
        self.assertIsNotNone(inbound_edit_block_reason(doc, other, self._today(), False))  # 非本人
        self.assertIsNone(inbound_edit_block_reason(doc, other, self._today(), True))       # 管理员
        old = create_and_post_inbound(company=self.c1, user=self.u, doc_date=date(2026, 1, 1),
            lines=[{"product": self.p, "quantity": Decimal("1"), "unit_price": Decimal("1")}])
        self.assertIn("跨月", inbound_edit_block_reason(old, self.u, self._today(), True))


class InboundListFilterTests(TestCase):
    """FilteredListMixin（#9）：入库单列表按日期区间 + 关键字筛选。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from apps.masterdata.models import Supplier
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="环氧树脂")
        cls.s1 = Supplier.objects.create(company=cls.c1, code="SUP1", name="甲供应商")
        cls.s2 = Supplier.objects.create(company=cls.c1, code="SUP2", name="乙供应商")
        create_and_post_inbound(company=cls.c1, user=None, doc_date=date(2026, 5, 1), supplier=cls.s1,
            lines=[{"product": cls.p, "quantity": Decimal("1"), "unit_price": Decimal("1")}])
        create_and_post_inbound(company=cls.c1, user=None, doc_date=date(2026, 6, 20), supplier=cls.s2,
            lines=[{"product": cls.p, "quantity": Decimal("1"), "unit_price": Decimal("1")}])
        U = get_user_model()
        cls.user = U.objects.create_superuser(username="root", password="x")

    def setUp(self):
        self.client.force_login(self.user)

    def _docnos(self, **params):
        resp = self.client.get("/purchasing/inbound/", params, SERVER_NAME="localhost")
        self.assertEqual(resp.status_code, 200)
        return {d.doc_no for d in resp.context["docs"]}

    def test_date_range_filters(self):
        got = self._docnos(**{"from": "2026-06-01", "to": "2026-06-30"})
        self.assertEqual(got, {"RK-C1-20260620-001"})

    def test_keyword_filters_by_supplier(self):
        got = self._docnos(q="乙供应商")
        self.assertEqual(got, {"RK-C1-20260620-001"})

    def test_no_filter_returns_all(self):
        self.assertEqual(len(self._docnos()), 2)

    def test_export_xlsx_respects_filter(self):
        from io import BytesIO
        from openpyxl import load_workbook
        resp = self.client.get("/purchasing/inbound/", {"q": "乙供应商", "export": "xlsx"},
                               SERVER_NAME="localhost")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # 第1行=公司抬头，第2行=报表名
        self.assertEqual(rows[1][0], "采购入库")           # 报表名行
        # 表头行（带样式后位于抬头+报表名+元信息之后）
        hdr_idx = next(i for i, r in enumerate(rows) if r[0] == "单据编号")
        data = rows[hdr_idx + 1:]
        self.assertEqual(len(data), 1)                     # 筛选后仅 1 行
        self.assertEqual(data[0][0], "RK-C1-20260620-001")


class InboundDeleteTests(TestCase):
    """采购入库硬删除（安全条件下）：反冲库存、彻底移除；非安全场景拦截。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="货A")
        cls.p2 = Product.objects.create(company=cls.c1, code="P002", name="货B")
        cls.u = get_user_model().objects.create_user(
            username="op", password="x", can_view_all_companies=True)

    def _today(self):
        from django.utils import timezone
        return timezone.localdate()

    def _inbound(self, product=None, qty="10", price="5"):
        return create_and_post_inbound(
            company=self.c1, user=self.u, doc_date=self._today(),
            lines=[{"product": product or self.p, "quantity": Decimal(qty),
                    "unit_price": Decimal(price)}])

    def test_delete_latest_inbound_reverses_stock(self):
        from apps.purchasing.services import delete_purchase_inbound
        from apps.purchasing.models import PurchaseInbound
        from apps.inventory.models import StockMove
        doc = self._inbound(qty="10", price="5")
        self.assertEqual(StockBalance.objects.get(company=self.c1, product=self.p).quantity,
                         Decimal("10.000"))
        delete_purchase_inbound(doc, user=self.u, today=self._today(), is_manager=False)
        self.assertEqual(PurchaseInbound.objects.filter(pk=doc.pk).count(), 0)   # 单据没了
        bal = StockBalance.objects.get(company=self.c1, product=self.p)
        self.assertEqual(bal.quantity, Decimal("0.000"))                          # 库存反冲
        self.assertEqual(bal.amount, Decimal("0.00"))
        # 该商品不留任何流水（含原始与反冲）
        self.assertFalse(StockMove.objects.filter(company=self.c1, product=self.p).exists())

    def test_delete_blocked_when_later_movement_exists(self):
        from apps.purchasing.services import delete_purchase_inbound, inbound_delete_block_reason
        first = self._inbound(qty="10", price="5")
        self._inbound(qty="5", price="6")   # 同商品后续又入库 → first 不再是最后一笔
        self.assertIsNotNone(inbound_delete_block_reason(first, self.u, self._today(), False))
        with self.assertRaises(Exception):
            delete_purchase_inbound(first, user=self.u, today=self._today(), is_manager=False)

    def test_other_product_later_movement_does_not_block(self):
        from apps.purchasing.services import inbound_delete_block_reason
        doc = self._inbound(product=self.p, qty="10", price="5")
        self._inbound(product=self.p2, qty="3", price="2")   # 别的商品，不影响
        self.assertIsNone(inbound_delete_block_reason(doc, self.u, self._today(), False))

    def test_delete_blocked_when_invoiced(self):
        from apps.finance.services import create_purchase_invoice
        from apps.masterdata.models import Supplier
        from apps.purchasing.services import inbound_delete_block_reason
        sup = Supplier.objects.create(company=self.c1, code="S1", name="供应商甲")
        doc = create_and_post_inbound(
            company=self.c1, user=self.u, doc_date=self._today(), supplier=sup,
            lines=[{"product": self.p, "quantity": Decimal("10"), "unit_price": Decimal("5")}])
        ln = doc.lines.first()
        create_purchase_invoice(
            company=self.c1, user=self.u, doc_date=self._today(), supplier=sup,
            lines=[{"product": self.p, "description": "", "amount_untaxed": Decimal("50"),
                    "tax_rate": Decimal("0"), "source_inbound_line": ln}])
        self.assertIn("发票", inbound_delete_block_reason(doc, self.u, self._today(), False))

    def test_delete_view(self):
        from apps.purchasing.models import PurchaseInbound
        doc = self._inbound()
        self.client.force_login(self.u)
        from django.contrib.auth.models import Permission
        for code in ("add_purchaseinbound", "view_purchaseinbound"):
            self.u.user_permissions.add(
                Permission.objects.get(content_type__app_label="purchasing", codename=code))
        r = self.client.post(f"/purchasing/inbound/{doc.pk}/delete/",
                             SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)
        self.assertEqual(PurchaseInbound.objects.filter(pk=doc.pk).count(), 0)


class InboundListTotalsTests(TestCase):
    """采购入库列表：不含税合计列 + 底部合计行。"""

    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        from apps.purchasing.services import create_and_post_inbound
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="货A")
        for _ in range(2):
            create_and_post_inbound(company=cls.c1, user=None, doc_date=date(2026, 6, 5),
                lines=[{"product": cls.p, "quantity": Decimal("10"),
                        "tax_inclusive_price": Decimal("113"), "tax_rate": Decimal("0.13")}])
        U = get_user_model()
        cls.user = U.objects.create_user(username="buy", password="x", can_view_all_companies=True)
        cls.user.user_permissions.add(
            Permission.objects.get(content_type__app_label="purchasing",
                                   codename="view_purchaseinbound"))

    def test_list_shows_untaxed_and_totals(self):
        self.client.force_login(self.user)
        resp = self.client.get("/purchasing/inbound/", SERVER_NAME="localhost")
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "不含税合计")
        t = resp.context["totals"]
        self.assertEqual(t["amount"], Decimal("2000.00"))    # 入库成本 2 × 1000
        self.assertEqual(t["untaxed"], Decimal("2000.00"))   # 2 × 1000
        self.assertEqual(t["taxed"], Decimal("2260.00"))     # 2 × 1130
        self.assertContains(resp, "合计")
