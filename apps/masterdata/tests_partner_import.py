"""往来单位 Excel 导入测试。"""

from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from apps.core.models import Company
from apps.masterdata.excel import (
    HEADERS,
    build_partner_template,
    import_partners,
    parse_partners_xlsx,
)
from apps.masterdata.models import BusinessPartner

User = get_user_model()


def _xlsx_bytes(rows):
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class PartnerExcelTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.c2 = Company.objects.create(code="C2", name="恒本源", short_name="恒本源")
        cls.user = User.objects.create_user(username="fin", password="x", is_superuser=True)
        cls.user.can_view_all_companies = True
        cls.user.save()

    def test_template_has_headers(self):
        data = build_partner_template()
        ws = load_workbook(BytesIO(data), read_only=True).active
        header = next(ws.iter_rows(values_only=True))
        self.assertEqual(list(header), HEADERS)

    def test_import_create_and_update(self):
        raw = _xlsx_bytes([
            ["W0001", "外部甲", "是", "否", "张三", "138", "", "", "", "是", ""],
            ["W0002", "恒本源", "是", "是", "", "", "", "", "C2", "是", "关联"],
        ])
        rows, errs = parse_partners_xlsx(BytesIO(raw))
        self.assertEqual(errs, [])
        created, updated, ierrs = import_partners(
            company=self.c1, user=self.user, rows=rows)
        self.assertEqual((created, updated, ierrs), (2, 0, []))
        a = BusinessPartner.objects.get(company=self.c1, code="W0001")
        self.assertTrue(a.is_customer and not a.is_supplier)
        b = BusinessPartner.objects.get(company=self.c1, code="W0002")
        self.assertEqual(b.related_company_id, self.c2.pk)

        raw2 = _xlsx_bytes([
            ["W0001", "外部甲改名", "是", "是", "", "", "", "", "", "是", "兼营"],
        ])
        rows2, _ = parse_partners_xlsx(BytesIO(raw2))
        created, updated, _ = import_partners(
            company=self.c1, user=self.user, rows=rows2)
        self.assertEqual((created, updated), (0, 1))
        a.refresh_from_db()
        self.assertEqual(a.name, "外部甲改名")
        self.assertTrue(a.is_supplier)

    def test_reject_self_related_company(self):
        raw = _xlsx_bytes([
            ["W0009", "自己", "是", "否", "", "", "", "", "C1", "是", ""],
        ])
        rows, _ = parse_partners_xlsx(BytesIO(raw))
        created, updated, errs = import_partners(
            company=self.c1, user=self.user, rows=rows)
        self.assertEqual((created, updated), (0, 0))
        self.assertTrue(any("不能选本公司" in e for e in errs))

    def test_import_view_requires_perm_and_works(self):
        self.client.force_login(self.user)
        r = self.client.get(reverse("partner_list"))
        self.assertContains(r, "导入")
        self.assertEqual(self.client.get(reverse("partner_template")).status_code, 200)

        raw = _xlsx_bytes([
            ["W0010", "网页导入", "否", "是", "", "", "", "", "", "是", ""],
        ])
        upload = SimpleUploadedFile(
            "p.xlsx", raw,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp = self.client.post(reverse("partner_import"), {"file": upload})
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(
            BusinessPartner.objects.filter(company=self.c1, code="W0010", is_supplier=True).exists())

    def test_import_forbidden_without_perm(self):
        u = User.objects.create_user(username="ro", password="x")
        u.can_view_all_companies = True
        u.save()
        # 仅 view，无 add
        u.user_permissions.add(Permission.objects.get(
            content_type__app_label="masterdata", codename="view_businesspartner"))
        self.client.force_login(u)
        self.assertEqual(self.client.get(reverse("partner_import")).status_code, 403)
