"""往来单位 Excel 导入 / 模板（openpyxl）。

列（首行表头）：
  编码 | 名称 | 客户 | 供应商 | 联系人 | 电话 | 税号 | 地址 | 对应关联企业 | 启用 | 备注

- 编码建议 `W` + 四位数字（如 W0001），公司内唯一。
- 客户/供应商/启用：是/否、1/0、TRUE/FALSE、Y/N（空=否；启用空=是）。
- 对应关联企业：填公司编号（C1/C2/C3）或简称/全称；外部单位留空。
- 同编码已存在则更新；至少勾选客户或供应商之一。
"""

from io import BytesIO

from openpyxl import Workbook, load_workbook

HEADERS = [
    "编码", "名称", "客户", "供应商", "联系人", "电话", "税号",
    "地址", "对应关联企业", "启用", "备注",
]


def _cell(row, idx, default=""):
    if idx >= len(row):
        return default
    v = row[idx]
    if v is None:
        return default
    return str(v).strip()


def _truthy(s, *, default=False) -> bool:
    if s is None or str(s).strip() == "":
        return default
    return str(s).strip().lower() in {"1", "y", "yes", "true", "是", "√", "x"}


def build_partner_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "往来单位"
    ws.append(HEADERS)
    ws.append([
        "W0001", "示例外部客户", "是", "否", "张三", "13800000000", "",
        "", "", "是", "",
    ])
    ws.append([
        "W0002", "示例兼营单位", "是", "是", "", "", "",
        "", "", "是", "客户+供应商都勾",
    ])
    ws.append([
        "W0003", "恒本源（关联）", "是", "是", "", "", "",
        "", "C2", "是", "对应关联企业填 C1/C2/C3",
    ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_partners_xlsx(upload) -> tuple[list[dict], list[str]]:
    """解析上传文件 → (rows, errors)。rows 为规范化 dict。"""
    wb = load_workbook(upload, data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], ["空文件"]
    # 允许表头顺序与 HEADERS 一致；若首格不是「编码」仍按固定列序读
    parsed, errors = [], []
    for i, raw in enumerate(rows_iter, start=2):
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        code = _cell(raw, 0)
        name = _cell(raw, 1)
        if not code and not name:
            continue
        if not code:
            errors.append(f"第{i}行：缺少编码")
            continue
        if not name:
            errors.append(f"第{i}行：缺少名称")
            continue
        is_customer = _truthy(_cell(raw, 2))
        is_supplier = _truthy(_cell(raw, 3))
        if not is_customer and not is_supplier:
            errors.append(f"第{i}行（{code}）：客户/供应商至少勾选一个（填「是」）")
            continue
        parsed.append({
            "row": i,
            "code": code,
            "name": name,
            "is_customer": is_customer,
            "is_supplier": is_supplier,
            "contact": _cell(raw, 4),
            "phone": _cell(raw, 5),
            "tax_no": _cell(raw, 6),
            "address": _cell(raw, 7),
            "related_company_key": _cell(raw, 8),
            "is_active": _truthy(_cell(raw, 9), default=True),
            "remark": _cell(raw, 10),
        })
    return parsed, errors


def resolve_related_company(company, key):
    """按编号/简称/全称匹配系统内公司；空则 None。不可挂本账套自己。"""
    from apps.core.models import Company

    if not key:
        return None, None
    qs = Company.objects.filter(is_active=True)
    hit = (qs.filter(code__iexact=key).first()
           or qs.filter(short_name=key).first()
           or qs.filter(name=key).first()
           or qs.filter(full_name=key).first())
    if hit is None:
        return None, f"找不到关联企业「{key}」"
    if hit.pk == company.pk:
        return None, "对应关联企业不能选本公司"
    return hit, None


def import_partners(*, company, user, rows) -> tuple[int, int, list[str]]:
    """按编码 upsert。返回 (created, updated, errors)。"""
    from django.db import transaction

    from .models import BusinessPartner

    created = updated = 0
    errors = []
    with transaction.atomic():
        for r in rows:
            related, err = resolve_related_company(company, r["related_company_key"])
            if err:
                errors.append(f"第{r['row']}行（{r['code']}）：{err}")
                continue
            defaults = {
                "name": r["name"],
                "contact": r["contact"],
                "phone": r["phone"],
                "tax_no": r["tax_no"],
                "address": r["address"],
                "related_company": related,
                "is_customer": r["is_customer"],
                "is_supplier": r["is_supplier"],
                "is_active": r["is_active"],
                "remark": r["remark"],
            }
            obj = BusinessPartner.objects.filter(company=company, code=r["code"]).first()
            if obj is None:
                BusinessPartner.objects.create(
                    company=company, created_by=user, code=r["code"], **defaults)
                created += 1
            else:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                updated += 1
    return created, updated, errors
