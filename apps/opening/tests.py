"""期初导入测试（M5-1）。"""

from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from openpyxl import Workbook

from apps.core.models import Company
from apps.finance.models import BankAccount, PurchaseInvoice
from apps.inventory.models import StockBalance
from apps.masterdata.models import Product, Supplier
from apps.opening import imports


def _xlsx(rows):
    wb = Workbook(); ws = wb.active
    for r in rows:
        ws.append(r)
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf


class OpeningImportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="环氧树脂")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户")

    def test_import_stock(self):
        # 数量金额式：数量 + 金额（不录单价）
        f = _xlsx([["商品编码", "数量", "金额"], ["P001", 100, 1000]])
        created, updated, skipped, errors = imports.import_stock(self.c1, None, f)
        self.assertEqual((created, updated, errors), (1, 0, []))
        bal = StockBalance.objects.get(company=self.c1, product=self.p)
        self.assertEqual(bal.quantity, Decimal("100.000"))
        self.assertEqual(bal.amount, Decimal("1000.00"))
        self.assertEqual(bal.avg_price, Decimal("10.00"))   # 由 金额/数量 得出
        # 未启用时再次导入覆盖更新（非跳过）
        f2 = _xlsx([["商品编码", "数量", "金额"], ["P001", 80, 800]])
        c2, u2, s2, _ = imports.import_stock(self.c1, None, f2, replace_existing=True)
        self.assertEqual((c2, u2, s2), (0, 1, 0))
        bal.refresh_from_db()
        self.assertEqual(bal.quantity, Decimal("80.000"))
        self.assertEqual(bal.amount, Decimal("800.00"))

    def test_import_combined_workbook(self):
        from openpyxl import Workbook
        from apps.finance.models import BankAccount as BA
        wb = Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet("期初库存"); ws.append(["商品编码", "数量", "金额"]); ws.append(["P001", 50, 750])
        ws2 = wb.create_sheet("期初银行存款"); ws2.append(["银行账户名称", "期初余额"]); ws2.append(["基本户", 8888])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        results = imports.import_combined(self.c1, None, buf)
        by = {r["kind"]: r for r in results}
        self.assertEqual(by["stock"]["created"], 1)
        self.assertEqual(by["bank"]["updated"], 1)
        self.assertEqual(StockBalance.objects.get(company=self.c1, product=self.p).amount, Decimal("750.00"))
        self.acc.refresh_from_db()
        self.assertEqual(self.acc.opening_balance, Decimal("8888.00"))

    def test_import_payable_creates_opening_invoice(self):
        f = _xlsx([["供应商编码", "期初应付金额"], ["S1", "5000"]])
        created, updated, skipped, errors = imports.import_payable(self.c1, None, f)
        self.assertEqual((created, updated, errors), (1, 0, []))
        inv = PurchaseInvoice.objects.get(company=self.c1, supplier=self.sup, is_opening=True)
        self.assertEqual(inv.outstanding, Decimal("5000.00"))
        # 覆盖更新
        f2 = _xlsx([["供应商编码", "期初应付金额"], ["S1", "4800"]])
        c2, u2, s2, _ = imports.import_payable(self.c1, None, f2, replace_existing=True)
        self.assertEqual((c2, u2, s2), (0, 1, 0))
        inv.refresh_from_db()
        self.assertEqual(inv.outstanding, Decimal("4800.00"))

    def test_import_bank_sets_opening_balance(self):
        f = _xlsx([["银行账户名称", "期初余额"], ["基本户", "12345.67"]])
        created, updated, skipped, errors = imports.import_bank(self.c1, None, f)
        self.assertEqual((updated, errors), (1, []))
        self.acc.refresh_from_db()
        self.assertEqual(self.acc.opening_balance, Decimal("12345.67"))

    def test_unknown_code_reports_error(self):
        f = _xlsx([["商品编码", "数量", "单价"], ["NOPE", 1, 1]])
        created, updated, skipped, errors = imports.import_stock(self.c1, None, f)
        self.assertEqual(created, 0)
        self.assertEqual(len(errors), 1)


class OverviewReportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from datetime import date
        from apps.inventory.services import post_inbound
        from apps.purchasing.services import create_and_post_inbound
        from apps.sales.services import create_and_post_outbound
        from apps.finance.services import (
            create_opening_payable, create_payment, create_purchase_invoice,
        )
        from apps.masterdata.models import Customer
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="货A")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户",
                                             opening_balance=Decimal("1000"))
        d = date(2026, 6, 1)
        # 期初库存 50@10=500
        post_inbound(cls.c1, cls.p, Decimal("50"), Decimal("10"),
                     source_type="Opening", source_no="期初")
        # 本期入库 100@10=1000，出库 30（成本 10 → 300）
        create_and_post_inbound(company=cls.c1, user=None, doc_date=d,
            lines=[{"product": cls.p, "quantity": Decimal("100"), "unit_price": Decimal("10")}])
        create_and_post_outbound(company=cls.c1, user=None, doc_date=d,
            lines=[{"product": cls.p, "quantity": Decimal("30")}])
        # 期初应付 2000；本期采购发票 1130；付款 500 核销
        create_opening_payable(company=cls.c1, user=None, supplier=cls.sup,
                               amount=Decimal("2000"), doc_date=d)
        inv = create_purchase_invoice(company=cls.c1, user=None, doc_date=d, supplier=cls.sup,
            lines=[{"product": cls.p, "description": "", "amount_untaxed": Decimal("1000"),
                    "tax_rate": Decimal("0.13")}])
        pay = create_payment(company=cls.c1, user=None, doc_date=d, bank_account=cls.acc,
                             supplier=cls.sup, amount=Decimal("500"))
        from apps.finance.services import allocate_payment
        allocate_payment(payment=pay, allocations=[{"invoice": inv, "amount": Decimal("500")}])

    def test_overview_reconciles_full_range(self):
        from datetime import date
        from apps.opening.reports import company_overview
        ov = company_overview(self.c1, date(2026, 1, 1), date(2030, 1, 1))  # 覆盖全部
        for key in ("bank", "stock", "payable", "receivable", "note_recv"):
            r = ov[key]
            self.assertEqual(r["opening"] + r["income"] - r["outgo"], r["ending"],
                             f"{key} 四列不勾稽")
        # 期末与区间无关：库存1200、银行500、应付2630
        self.assertEqual(ov["stock"]["ending"], Decimal("1200.00"))
        self.assertEqual(ov["bank"]["ending"], Decimal("500.00"))
        self.assertEqual(ov["payable"]["ending"], Decimal("2630.00"))

    def test_future_range_moves_all_to_opening(self):
        from datetime import date
        from apps.opening.reports import company_overview
        ov = company_overview(self.c1, date(2030, 1, 1), date(2030, 12, 31))  # 全在区间前
        # 区间内无活动 → 收入/发出为0，期初=期末
        self.assertEqual(ov["stock"]["income"], Decimal("0.00"))
        self.assertEqual(ov["stock"]["outgo"], Decimal("0.00"))
        self.assertEqual(ov["stock"]["opening"], ov["stock"]["ending"])


class AccountBalanceTableTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from datetime import date
        from apps.inventory.services import post_inbound
        from apps.purchasing.services import create_and_post_inbound
        from apps.finance.services import (
            allocate_payment, create_payment, create_purchase_invoice,
        )
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="货A")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.acc = BankAccount.objects.create(company=cls.c1, name="基本户",
                                             opening_balance=Decimal("1000"))
        d = date(2026, 6, 1)
        post_inbound(cls.c1, cls.p, Decimal("50"), Decimal("10"),
                     source_type="Opening", source_no="期初")
        create_and_post_inbound(company=cls.c1, user=None, doc_date=d,
            lines=[{"product": cls.p, "quantity": Decimal("100"), "unit_price": Decimal("10")}])
        inv = create_purchase_invoice(company=cls.c1, user=None, doc_date=d, supplier=cls.sup,
            lines=[{"product": cls.p, "description": "", "amount_untaxed": Decimal("1000"),
                    "tax_rate": Decimal("0.13")}])  # 应付 1130
        pay = create_payment(company=cls.c1, user=None, doc_date=d, bank_account=cls.acc,
                             supplier=cls.sup, amount=Decimal("500"))
        allocate_payment(payment=pay, allocations=[{"invoice": inv, "amount": Decimal("500")}])

    def test_account_balance_detail_rows(self):
        from datetime import date
        from apps.opening.reports import account_balance_table
        t = account_balance_table([self.c1], date(2026, 1, 1), date(2030, 1, 1))
        # 银行：期初1000 + 付款500流出 → 期末500
        bank = t["bank"][0]
        self.assertEqual(bank["ending"], Decimal("500.00"))
        self.assertEqual(bank["outgo"], Decimal("500.00"))
        # 库存：50@10 + 100@10 = 1500
        stock = t["stock"][0]
        self.assertEqual(stock["ending"], Decimal("1500.00"))
        # 应付：发票1130 - 付款核销500 = 630
        ap = t["payable"][0]
        self.assertEqual(ap["ending"], Decimal("630.00"))
        for sec in (bank, stock, ap):
            self.assertEqual(sec["opening"] + sec["income"] - sec["outgo"], sec["ending"])

    def _login_fin(self):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        user = get_user_model().objects.create_user(
            username="fin", password="x", can_view_all_companies=True)
        user.user_permissions.add(
            Permission.objects.get(content_type__app_label="finance", codename="view_bankjournal"))
        self.client.force_login(user)

    def test_account_balance_page_section_totals(self):
        """每个分组（银行/应收/应付/库存）底部有「合计」行，数值=各行之和。"""
        self._login_fin()
        resp = self.client.get("/account-balance/?from=2026-01-01&to=2030-01-01",
                               SERVER_NAME="localhost")
        self.assertEqual(resp.status_code, 200)
        blocks = {b["key"]: b for b in resp.context["blocks"]}
        # 银行合计 = 唯一账户期末 500；应付合计 = 630
        self.assertEqual(blocks["bank"]["total"]["ending"], Decimal("500.00"))
        self.assertEqual(blocks["payable"]["total"]["ending"], Decimal("630.00"))
        self.assertEqual(blocks["stock"]["total"]["ending"], Decimal("1500.00"))
        for b in blocks.values():
            t = b["total"]
            self.assertEqual(t["opening"] + t["income"] - t["outgo"], t["ending"])
        self.assertContains(resp, "合计")

    def test_account_balance_export_has_total_rows(self):
        """导出 Excel 在每个有数据的分组后附「合计」行。"""
        from openpyxl import load_workbook
        self._login_fin()
        resp = self.client.get(
            "/account-balance/?from=2026-01-01&to=2030-01-01&export=xlsx",
            SERVER_NAME="localhost")
        self.assertEqual(resp.status_code, 200)
        ws = load_workbook(BytesIO(resp.content)).active
        rows = list(ws.iter_rows(values_only=True))
        total_rows = [r for r in rows if r[2] == "合计"]
        # 本数据有 银行/应付/库存 三个分组有数据（无应收）
        self.assertEqual(len(total_rows), 3)
        bank_total = next(r for r in total_rows if r[0] == "银行存款明细账户")
        self.assertEqual(Decimal(str(bank_total[6])), Decimal("500.00"))


class QueryCenterTests(TestCase):
    """查询中心（M11）：跨公司组合查询。"""

    @classmethod
    def setUpTestData(cls):
        from datetime import date
        from apps.inventory.services import post_inbound
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.c2 = Company.objects.create(code="C2", name="恒本源", short_name="恒本源")
        cls.p1 = Product.objects.create(company=cls.c1, code="P001", name="环氧树脂")
        cls.p2 = Product.objects.create(company=cls.c2, code="P001", name="环氧树脂")
        post_inbound(cls.c1, cls.p1, Decimal("10"), Decimal("5"), date=date(2026, 6, 5))   # 50
        post_inbound(cls.c2, cls.p2, Decimal("20"), Decimal("5"), date=date(2026, 6, 5))   # 100

    def _run(self, **params):
        from apps.opening.query import run_query
        comps = params.pop("companies", [self.c1, self.c2])
        from datetime import date
        return run_query(params.pop("subject", "stock_moves"), comps,
                         date(2026, 6, 1), date(2026, 6, 30),
                         {"q": params.get("q", ""), "direction": params.get("direction", ""),
                          "entry_type": "", "status": ""})

    def test_multi_company_aggregates(self):
        r = self._run()
        self.assertEqual(len(r["rows"]), 2)             # 两家各一条入库
        # 收入金额合计列(索引5) = 50 + 100 = 150
        self.assertEqual(r["totals"][5], Decimal("150.00"))

    def test_single_company_scope(self):
        r = self._run(companies=[self.c1])
        self.assertEqual(len(r["rows"]), 1)
        self.assertEqual(r["totals"][5], Decimal("50.00"))

    def test_keyword_and_direction(self):
        r = self._run(q="环氧")
        self.assertEqual(len(r["rows"]), 2)
        r2 = self._run(direction="out")                 # 无出库
        self.assertEqual(len(r2["rows"]), 0)


class ReconciliationTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from datetime import date
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        from apps.finance.services import create_purchase_invoice
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.sup = Supplier.objects.create(company=cls.c1, code="S1", name="供应商甲")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="货A")
        create_purchase_invoice(company=cls.c1, user=None, doc_date=date(2026, 6, 1), supplier=cls.sup,
            lines=[{"product": cls.p, "description": "", "amount_untaxed": Decimal("1000"),
                    "tax_rate": Decimal("0.13")}])  # 应付 1130
        U = get_user_model()
        cls.user = U.objects.create_user(username="fin", password="x", can_view_all_companies=True)
        cls.user.user_permissions.add(
            Permission.objects.get(content_type__app_label="finance", codename="view_purchaseinvoice"))

    def test_payable_recon_lines(self):
        from apps.opening.reports import recon_lines
        lines = recon_lines(self.c1, "payable")
        self.assertEqual(len(lines), 1)
        self.assertEqual(lines[0]["system_amount"], Decimal("1130.00"))

    def test_reconciliation_post_persists_diff(self):
        from apps.opening.models import ReconciliationRun
        self.client.force_login(self.user)
        # 外部值填 1100 → 差异 -30
        resp = self.client.post("/reconciliation/", {
            "category": "payable", "as_of": "2026-06-30", "ext-0": "1100",
        }, SERVER_NAME="localhost")
        self.assertEqual(resp.status_code, 200)
        run = ReconciliationRun.objects.get(company=self.c1, category="payable")
        line = run.lines.get()
        self.assertEqual(line.system_amount, Decimal("1130.00"))
        self.assertEqual(line.external_amount, Decimal("1100.00"))
        self.assertEqual(line.diff, Decimal("-30.00"))


class OpeningImportViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        from django.contrib.auth import get_user_model
        from django.contrib.auth.models import Permission
        cls.c1 = Company.objects.create(code="C1", name="安博诺", short_name="安博诺")
        cls.p = Product.objects.create(company=cls.c1, code="P001", name="货A")
        U = get_user_model()
        cls.user = U.objects.create_user(username="fin", password="x", can_view_all_companies=True)
        cls.user.user_permissions.add(
            Permission.objects.get(content_type__app_label="finance", codename="add_purchaseinvoice"))

    def test_import_view_uploads_combined(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from openpyxl import Workbook
        from apps.inventory.models import StockBalance
        wb = Workbook(); wb.remove(wb.active)
        ws = wb.create_sheet("期初库存"); ws.append(["商品编码", "数量", "金额"]); ws.append(["P001", 20, 100])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        self.client.force_login(self.user)
        resp = self.client.post("/opening/", {
            "file": SimpleUploadedFile("o.xlsx", buf.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        }, SERVER_NAME="localhost", follow=True)
        self.assertEqual(resp.status_code, 200)
        bal = StockBalance.objects.get(company=self.c1, product=self.p)
        self.assertEqual(bal.quantity, Decimal("20.000"))
        self.assertEqual(bal.amount, Decimal("100.00"))

    def test_combined_template_download(self):
        from openpyxl import load_workbook
        self.client.force_login(self.user)
        resp = self.client.get("/opening/template/all/", SERVER_NAME="localhost")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        wb = load_workbook(BytesIO(resp.content))
        self.assertIn("期初库存", wb.sheetnames)
        self.assertIn("期初应付", wb.sheetnames)
        self.assertEqual([c.value for c in wb["期初库存"][1]], ["商品编码", "数量", "金额"])

    def test_clear_kind_payable_when_stock_has_biz(self):
        """已有日常库存业务时，整体清空被拦，但可分类清空期初应付。"""
        from apps.finance.services import create_opening_payable
        from apps.purchasing.services import create_and_post_inbound
        from apps.masterdata.models import Supplier
        from datetime import date
        sup = Supplier.objects.create(company=self.c1, code="S9", name="供应商九")
        create_opening_payable(company=self.c1, user=None, supplier=sup,
                               amount=Decimal("1000"), doc_date=date(2026, 6, 1))
        create_and_post_inbound(
            company=self.c1, user=None, doc_date=date(2026, 6, 15),
            lines=[{"product": self.p, "quantity": Decimal("1"), "unit_price": Decimal("10")}])
        self.client.force_login(self.user)
        resp = self.client.post("/opening/clear/payable/", SERVER_NAME="localhost", follow=True)
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(PurchaseInvoice.objects.filter(company=self.c1, is_opening=True).exists())
