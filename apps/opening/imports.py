"""期初数据 Excel 导入（M5-1，SPEC §8.1）。

8 类期初合并在**一个工作簿、每类一个 sheet**：下载一个模板→各 sheet 填好→上传一次全部导入。
导入到启用日 settings.OPENING_DATE。
- 期初未启用（opening_locked=False）：重复导入**覆盖更新**已有期初行（按商品/往来/票号匹配）。
- 期初已锁定：不可导入。
库存为「数量金额式」（数量 + 金额，不录单价，均价由系统按金额/数量得出）。
发出商品 / 应付暂估：只挂台账（标记 is_opening），**不重复加减库存**（库存期初已反映仓库结存）。
"""

from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.conf import settings
from openpyxl import Workbook, load_workbook

from apps.finance.models import BankAccount, NotePayable, NoteReceivable, PurchaseInvoice, SalesInvoice
from apps.finance.services import (
    create_note_payable,
    create_note_receivable,
    create_opening_payable,
    create_opening_receivable,
)
from apps.core.docnum import next_doc_no
from apps.core.money import ZERO_MONEY, round_money, round_qty
from apps.inventory.models import StockMove
from apps.inventory.services import post_inbound
from apps.masterdata.models import Customer, Product, Supplier
from apps.purchasing.models import PurchaseInbound, PurchaseInboundLine
from apps.sales.models import SalesOutbound, SalesOutboundLine

OPENING = settings.OPENING_DATE

# 各类期初的表头（首行）
TEMPLATES = {
    "stock": ["商品编码", "数量", "金额"],
    "payable": ["供应商编码", "期初应付金额"],
    "receivable": ["客户编码", "期初应收金额"],
    "bank": ["银行账户名称", "期初余额"],
    "note_receivable": ["票据号", "金额", "来源客户编码(可空)", "到期日(可空)"],
    "note_payable": ["票据号", "金额", "收票供应商编码", "到期日(可空)"],
    "goods_shipped": ["客户编码(可空)", "商品编码", "数量", "结转成本", "售价不含税(可空)"],
    "ap_accrual": ["供应商编码", "商品编码", "数量", "不含税金额"],
}

# 合并工作簿里各 sheet 的标题（也用于导入时按标题定位）
SHEET_TITLES = {
    "stock": "期初库存",
    "payable": "期初应付",
    "receivable": "期初应收",
    "bank": "期初银行存款",
    "note_receivable": "期初应收票据",
    "note_payable": "期初应付票据",
    "goods_shipped": "期初发出商品",
    "ap_accrual": "期初应付账款-暂估",
}

_ALL_HEADERS = sum(TEMPLATES.values(), [])


def build_template(kind) -> bytes:
    """单类模板（保留，供测试/旧链接用）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "期初"
    ws.append(TEMPLATES[kind])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_combined_template() -> bytes:
    """一个工作簿、每类一个 sheet 的合并模板。"""
    wb = Workbook()
    wb.remove(wb.active)  # 去掉默认空 sheet
    for kind, title in SHEET_TITLES.items():
        ws = wb.create_sheet(title=title)
        ws.append(TEMPLATES[kind])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _rows_ws(ws):
    out = []
    for idx, values in enumerate(ws.iter_rows(values_only=True), start=1):
        values = list(values)
        if not any(v not in (None, "") for v in values):
            continue
        first = str(values[0]).strip() if values and values[0] is not None else ""
        if idx == 1 or first in _ALL_HEADERS:  # 跳过表头
            continue
        out.append((idx, values))
    return out


def _rows(file):
    """读取上传文件的活动 sheet（单类导入用）。"""
    wb = load_workbook(file, read_only=True, data_only=True)
    return _rows_ws(wb.active)


def _dec(v):
    """把 Excel 单元格值转换为 Decimal。

    兼容：
    - 数字单元格（int/float/Decimal）
    - 文本数字（含千分位逗号，如 366,292.03）
    """
    try:
        if v in (None, ""):
            return None
        s = str(v).strip()
        # Excel/人工常见：千分位分隔符，先去掉再转
        s = s.replace(",", "")
        return Decimal(s)
    except (InvalidOperation, ValueError, TypeError):
        return None


def _date(v):
    from datetime import date, datetime
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    if v in (None, ""):
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _apply_stock(company, user, rows, *, replace_existing=False):
    """期初库存（数量金额式）：数量 + 金额，均价由系统按 金额/数量 得出。

    replace_existing=True（期初未锁定时）：已有期初流水则尝试覆盖——仅当该商品
    尚无后续非期初出入库时允许（否则移动加权链已引用，须逐笔调整）。
    """
    from apps.inventory.services import post_inbound, reverse_move

    created = updated = skipped = 0
    errors = []
    for idx, vals in rows:
        code = str(vals[0] or "").strip()
        qty = _dec(vals[1] if len(vals) > 1 else None)
        amount = _dec(vals[2] if len(vals) > 2 else None)
        product = Product.objects.filter(company=company, code=code).first()
        if not product:
            errors.append(f"第{idx}行：商品编码「{code}」不存在"); continue
        if qty is None or amount is None or qty == 0:
            errors.append(f"第{idx}行：数量/金额无效（数量不能为 0、金额必填）"); continue
        existing = StockMove.objects.filter(
            company=company, product=product, source_type="Opening")
        if existing.exists():
            if not replace_existing:
                skipped += 1; continue
            if StockMove.objects.filter(company=company, product=product).exclude(
                    source_type="Opening").exists():
                errors.append(
                    f"第{idx}行：商品「{code}」已有后续出入库，不能通过导入改期初（请逐笔调整）")
                continue
            for m in list(existing):
                reverse_move(m, date=OPENING, source_type="OpeningReplace", source_no="期初修正")
            updated += 1
        else:
            created += 1
        post_inbound(company, product, qty, ZERO_MONEY, amount=amount, date=OPENING,
                     source_type="Opening", source_no="期初", allow_nonpositive=True)
    return created, updated, skipped, errors


def _apply_payable(company, user, rows, *, replace_existing=False):
    created = updated = skipped = 0
    errors = []
    for idx, vals in rows:
        code = str(vals[0] or "").strip()
        amount = _dec(vals[1] if len(vals) > 1 else None)
        sup = Supplier.objects.filter(company=company, code=code).first()
        if not sup:
            errors.append(f"第{idx}行：供应商编码「{code}」不存在"); continue
        if amount is None:
            errors.append(f"第{idx}行：金额无效"); continue
        if amount == 0:
            skipped += 1; continue   # 0 视为空行跳过；负数允许（红字期初）
        inv = PurchaseInvoice.objects.filter(
            company=company, supplier=sup, is_opening=True).first()
        if inv:
            if not replace_existing:
                skipped += 1; continue
            if inv.settled_amount > 0:
                errors.append(f"第{idx}行：供应商「{code}」期初应付已有核销，不能覆盖")
                continue
            inv.amount_untaxed = amount
            inv.tax_amount = ZERO_MONEY
            inv.amount_taxed = amount
            inv.save(update_fields=["amount_untaxed", "tax_amount", "amount_taxed"])
            ln = inv.lines.first()
            if ln:
                ln.amount_untaxed = amount
                ln.tax_amount = ZERO_MONEY
                ln.amount_taxed = amount
                ln.save(update_fields=["amount_untaxed", "tax_amount", "amount_taxed"])
            updated += 1
        else:
            create_opening_payable(company=company, user=user, supplier=sup, amount=amount, doc_date=OPENING)
            created += 1
    return created, updated, skipped, errors


def _apply_receivable(company, user, rows, *, replace_existing=False):
    created = updated = skipped = 0
    errors = []
    for idx, vals in rows:
        code = str(vals[0] or "").strip()
        amount = _dec(vals[1] if len(vals) > 1 else None)
        cust = Customer.objects.filter(company=company, code=code).first()
        if not cust:
            errors.append(f"第{idx}行：客户编码「{code}」不存在"); continue
        if amount is None:
            errors.append(f"第{idx}行：金额无效"); continue
        if amount == 0:
            skipped += 1; continue   # 0 视为空行跳过；负数允许（红字期初）
        inv = SalesInvoice.objects.filter(
            company=company, customer=cust, is_opening=True).first()
        if inv:
            if not replace_existing:
                skipped += 1; continue
            if inv.settled_amount > 0:
                errors.append(f"第{idx}行：客户「{code}」期初应收已有核销，不能覆盖")
                continue
            inv.amount_untaxed = amount
            inv.tax_amount = ZERO_MONEY
            inv.amount_taxed = amount
            inv.save(update_fields=["amount_untaxed", "tax_amount", "amount_taxed"])
            ln = inv.lines.first()
            if ln:
                ln.amount_untaxed = amount
                ln.tax_amount = ZERO_MONEY
                ln.amount_taxed = amount
                ln.save(update_fields=["amount_untaxed", "tax_amount", "amount_taxed"])
            updated += 1
        else:
            create_opening_receivable(company=company, user=user, customer=cust, amount=amount, doc_date=OPENING)
            created += 1
    return created, updated, skipped, errors


def _apply_bank(company, user, rows, *, replace_existing=False):
    updated = skipped = 0
    errors = []
    for idx, vals in rows:
        name = str(vals[0] or "").strip()
        bal = _dec(vals[1] if len(vals) > 1 else None)
        acc = BankAccount.objects.filter(company=company, name=name).first()
        if not acc:
            errors.append(f"第{idx}行：银行账户「{name}」不存在"); continue
        if bal is None:
            errors.append(f"第{idx}行：余额无效"); continue
        acc.opening_balance = bal
        acc.save(update_fields=["opening_balance"])
        updated += 1
    return 0, updated, skipped, errors


def _apply_note_receivable(company, user, rows, *, replace_existing=False):
    created = updated = skipped = 0
    errors = []
    for idx, vals in rows:
        note_no = str(vals[0] or "").strip()
        amount = _dec(vals[1] if len(vals) > 1 else None)
        cust = Customer.objects.filter(company=company, code=str(vals[2] or "").strip()).first() if len(vals) > 2 and vals[2] else None
        due = _date(vals[3]) if len(vals) > 3 else None
        if amount is None:
            errors.append(f"第{idx}行：金额无效"); continue
        if amount == 0:
            skipped += 1; continue   # 0 视为空行跳过；负数允许
        existing = (NoteReceivable.objects.filter(
            company=company, note_no=note_no, is_opening=True).first()
                    if note_no else None)
        if existing:
            if not replace_existing:
                skipped += 1; continue
            if existing.settled_amount > 0:
                errors.append(f"第{idx}行：票据「{note_no}」已有使用，不能覆盖")
                continue
            existing.amount = amount
            existing.customer = cust
            existing.due_date = due
            existing.save(update_fields=["amount", "customer", "due_date"])
            updated += 1
        elif note_no and NoteReceivable.objects.filter(company=company, note_no=note_no).exists():
            skipped += 1; continue
        else:
            create_note_receivable(company=company, user=user, draw_date=OPENING, amount=amount,
                                   customer=cust, note_no=note_no, due_date=due, is_opening=True)
            created += 1
    return created, updated, skipped, errors


def _apply_note_payable(company, user, rows, *, replace_existing=False):
    created = updated = skipped = 0
    errors = []
    for idx, vals in rows:
        note_no = str(vals[0] or "").strip()
        amount = _dec(vals[1] if len(vals) > 1 else None)
        sup = Supplier.objects.filter(company=company, code=str(vals[2] or "").strip()).first() if len(vals) > 2 and vals[2] else None
        due = _date(vals[3]) if len(vals) > 3 else None
        if amount is None:
            errors.append(f"第{idx}行：金额无效"); continue
        if amount == 0:
            skipped += 1; continue   # 0 视为空行跳过；负数允许
        if not sup:
            errors.append(f"第{idx}行：收票供应商编码无效"); continue
        existing = (NotePayable.objects.filter(
            company=company, note_no=note_no, is_opening=True).first()
                    if note_no else None)
        if existing:
            if not replace_existing:
                skipped += 1; continue
            if existing.settled_amount > 0:
                errors.append(f"第{idx}行：票据「{note_no}」已有使用，不能覆盖")
                continue
            existing.amount = amount
            existing.supplier = sup
            existing.due_date = due
            existing.save(update_fields=["amount", "supplier", "due_date"])
            updated += 1
        elif note_no and NotePayable.objects.filter(company=company, note_no=note_no).exists():
            skipped += 1; continue
        else:
            create_note_payable(company=company, user=user, draw_date=OPENING, amount=amount,
                                supplier=sup, note_no=note_no, due_date=due, is_opening=True)
            created += 1
    return created, updated, skipped, errors


def _find_opening_outbound_line(company, product, customer):
    """按 公司+商品+客户 匹配一条期初发出商品行。"""
    qs = (SalesOutboundLine.objects
          .filter(outbound__company=company, outbound__is_opening=True, product=product)
          .select_related("outbound"))
    if customer is None:
        qs = qs.filter(outbound__customer__isnull=True)
    else:
        qs = qs.filter(outbound__customer=customer)
    return qs.first()


def _apply_goods_shipped(company, user, rows, *, replace_existing=False):
    """期初发出商品：已出库未开票成本台账，不减库存。"""
    from django.db import transaction

    created = updated = skipped = 0
    errors = []
    for idx, vals in rows:
        cust_code = str(vals[0] or "").strip() if vals else ""
        code = str(vals[1] or "").strip() if len(vals) > 1 else ""
        qty = _dec(vals[2] if len(vals) > 2 else None)
        cost = _dec(vals[3] if len(vals) > 3 else None)
        untaxed = _dec(vals[4] if len(vals) > 4 else None)
        cust = None
        if cust_code:
            cust = Customer.objects.filter(company=company, code=cust_code).first()
            if not cust:
                errors.append(f"第{idx}行：客户编码「{cust_code}」不存在"); continue
        product = Product.objects.filter(company=company, code=code).first()
        if not product:
            errors.append(f"第{idx}行：商品编码「{code}」不存在"); continue
        if qty is None or cost is None or qty == 0:
            errors.append(f"第{idx}行：数量/结转成本无效（数量不能为 0）"); continue
        qty = round_qty(qty)
        cost = round_money(cost)
        if untaxed is None:
            untaxed = ZERO_MONEY
        else:
            untaxed = round_money(untaxed)
        unit_cost = round_money(cost / qty) if qty else ZERO_MONEY
        existing = _find_opening_outbound_line(company, product, cust)
        if existing:
            if not replace_existing:
                skipped += 1; continue
            from apps.finance.models import SalesInvoiceLine
            if SalesInvoiceLine.objects.filter(source_outbound_line=existing).exists():
                errors.append(f"第{idx}行：商品「{code}」期初发出商品已开票关联，不能覆盖")
                continue
            existing.quantity = qty
            existing.unit_cost = unit_cost
            existing.amount = cost
            existing.amount_untaxed = untaxed
            existing.tax_amount = ZERO_MONEY
            existing.amount_taxed = untaxed
            existing.sale_unit_price = round_money(untaxed / qty) if qty and untaxed else ZERO_MONEY
            existing.save()
            doc = existing.outbound
            doc.total_quantity = qty
            doc.total_cost = cost
            doc.total_untaxed = untaxed
            doc.total_tax = ZERO_MONEY
            doc.total_taxed = untaxed
            doc.save(update_fields=["total_quantity", "total_cost", "total_untaxed",
                                    "total_tax", "total_taxed"])
            updated += 1
            continue
        with transaction.atomic():
            doc = SalesOutbound.objects.create(
                company=company, created_by=user,
                doc_no=next_doc_no(SalesOutbound, company, "CK", OPENING),
                doc_date=OPENING, customer=cust,
                sales_type=SalesOutbound.SalesType.SALE,
                status=SalesOutbound.Status.POSTED,
                remark="期初发出商品", is_opening=True,
                total_quantity=qty, total_cost=cost,
                total_untaxed=untaxed, total_tax=ZERO_MONEY, total_taxed=untaxed,
            )
            SalesOutboundLine.objects.create(
                outbound=doc, product=product, quantity=qty,
                sale_unit_price=round_money(untaxed / qty) if qty and untaxed else ZERO_MONEY,
                amount_untaxed=untaxed, tax_amount=ZERO_MONEY, amount_taxed=untaxed,
                unit_cost=unit_cost, amount=cost, stock_move=None,
            )
            created += 1
    return created, updated, skipped, errors


def _find_opening_inbound_line(company, product, supplier):
    qs = (PurchaseInboundLine.objects
          .filter(inbound__company=company, inbound__is_opening=True, product=product,
                  inbound__supplier=supplier)
          .select_related("inbound"))
    return qs.first()


def _apply_ap_accrual(company, user, rows, *, replace_existing=False):
    """期初应付账款-暂估：已入库未收票（不含税），不加库存。"""
    from django.db import transaction

    created = updated = skipped = 0
    errors = []
    for idx, vals in rows:
        sup_code = str(vals[0] or "").strip()
        code = str(vals[1] or "").strip() if len(vals) > 1 else ""
        qty = _dec(vals[2] if len(vals) > 2 else None)
        untaxed = _dec(vals[3] if len(vals) > 3 else None)
        sup = Supplier.objects.filter(company=company, code=sup_code).first()
        if not sup:
            errors.append(f"第{idx}行：供应商编码「{sup_code}」不存在"); continue
        product = Product.objects.filter(company=company, code=code).first()
        if not product:
            errors.append(f"第{idx}行：商品编码「{code}」不存在"); continue
        if qty is None or untaxed is None or qty == 0:
            errors.append(f"第{idx}行：数量/不含税金额无效（数量不能为 0）"); continue
        qty = round_qty(qty)
        untaxed = round_money(untaxed)
        unit = round_money(untaxed / qty) if qty else ZERO_MONEY
        existing = _find_opening_inbound_line(company, product, sup)
        if existing:
            if not replace_existing:
                skipped += 1; continue
            from apps.finance.models import PurchaseInvoiceLine
            if PurchaseInvoiceLine.objects.filter(source_inbound_line=existing).exists():
                errors.append(f"第{idx}行：商品「{code}」期初暂估已收票关联，不能覆盖")
                continue
            existing.quantity = qty
            existing.unit_price = unit
            existing.amount_untaxed = untaxed
            existing.tax_amount = ZERO_MONEY
            existing.amount_taxed = untaxed
            existing.amount = untaxed
            existing.save()
            doc = existing.inbound
            doc.total_quantity = qty
            doc.total_amount = untaxed
            doc.total_untaxed = untaxed
            doc.total_tax = ZERO_MONEY
            doc.total_taxed = untaxed
            doc.save(update_fields=["total_quantity", "total_amount", "total_untaxed",
                                    "total_tax", "total_taxed"])
            updated += 1
            continue
        with transaction.atomic():
            doc = PurchaseInbound.objects.create(
                company=company, created_by=user,
                doc_no=next_doc_no(PurchaseInbound, company, "RK", OPENING),
                doc_date=OPENING, supplier=sup,
                purchase_type=PurchaseInbound.PurchaseType.EXTERNAL,
                status=PurchaseInbound.Status.POSTED,
                remark="期初应付账款-暂估", is_opening=True,
                total_quantity=qty, total_amount=untaxed,
                total_untaxed=untaxed, total_tax=ZERO_MONEY, total_taxed=untaxed,
            )
            PurchaseInboundLine.objects.create(
                inbound=doc, product=product, quantity=qty, unit_price=unit,
                amount_untaxed=untaxed, tax_amount=ZERO_MONEY, amount_taxed=untaxed,
                amount=untaxed, stock_move=None,
            )
            created += 1
    return created, updated, skipped, errors


_APPLY = {
    "stock": ("期初库存", _apply_stock),
    "payable": ("期初应付", _apply_payable),
    "receivable": ("期初应收", _apply_receivable),
    "bank": ("期初银行存款", _apply_bank),
    "note_receivable": ("期初应收票据", _apply_note_receivable),
    "note_payable": ("期初应付票据", _apply_note_payable),
    "goods_shipped": ("期初发出商品", _apply_goods_shipped),
    "ap_accrual": ("期初应付账款-暂估", _apply_ap_accrual),
}


def import_combined(company, user, file, *, replace_existing=False):
    """读取合并工作簿，按 sheet 标题逐类导入；返回 [{kind,label,created,updated,skipped,errors}]。"""
    wb = load_workbook(file, read_only=True, data_only=True)
    titles = set(wb.sheetnames)
    results = []
    for kind, title in SHEET_TITLES.items():
        if title not in titles:
            continue
        label, fn = _APPLY[kind]
        rows = _rows_ws(wb[title])
        created, updated, skipped, errors = fn(company, user, rows, replace_existing=replace_existing)
        results.append({"kind": kind, "label": label, "created": created, "updated": updated,
                        "skipped": skipped, "errors": errors})
    return results


# 单类导入（保留：旧链接/测试用）。file = 上传文件的活动 sheet。
def import_stock(company, user, file, *, replace_existing=False):
    return _apply_stock(company, user, _rows(file), replace_existing=replace_existing)


def import_payable(company, user, file, *, replace_existing=False):
    return _apply_payable(company, user, _rows(file), replace_existing=replace_existing)


def import_receivable(company, user, file, *, replace_existing=False):
    return _apply_receivable(company, user, _rows(file), replace_existing=replace_existing)


def import_bank(company, user, file, *, replace_existing=False):
    return _apply_bank(company, user, _rows(file), replace_existing=replace_existing)


def import_note_receivable(company, user, file, *, replace_existing=False):
    return _apply_note_receivable(company, user, _rows(file), replace_existing=replace_existing)


def import_note_payable(company, user, file, *, replace_existing=False):
    return _apply_note_payable(company, user, _rows(file), replace_existing=replace_existing)


IMPORTERS = {
    "stock": ("期初库存", import_stock),
    "payable": ("期初应付", import_payable),
    "receivable": ("期初应收", import_receivable),
    "bank": ("期初银行存款", import_bank),
    "note_receivable": ("期初应收票据", import_note_receivable),
    "note_payable": ("期初应付票据", import_note_payable),
    "goods_shipped": ("期初发出商品", lambda c, u, f, **kw: _apply_goods_shipped(c, u, _rows(f), **kw)),
    "ap_accrual": ("期初应付账款-暂估", lambda c, u, f, **kw: _apply_ap_accrual(c, u, _rows(f), **kw)),
}
